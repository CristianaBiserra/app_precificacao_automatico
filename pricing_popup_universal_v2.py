import os
import re
import csv
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_TITLE = "Simulador de Precificacao - Versao Universal"
DEFAULT_WORKBOOK = os.path.join(os.path.dirname(os.path.abspath(__file__)), "1 NOVO SIMULADOR_PRECIFICACAO_V2.xlsx")

PIS_COFINS_RATE = Decimal("0.0925")
FRETE_FOB_RATE_RN = Decimal("0.10")
FRETE_FOB_RATE_PE = Decimal("0.05")
ROUND_MONEY = Decimal("0.01")

CNPJ_RN = "18.217.682/0004-05"
CNPJ_PE = "18.217.682/0001-54"
CNPJ_RN_DIGITS = "18217682000405"
CNPJ_PE_DIGITS = "18217682000154"


def q2(value: Decimal) -> Decimal:
    return value.quantize(ROUND_MONEY, rounding=ROUND_HALF_UP)


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip().replace("R$", "").replace("%", "").replace(" ", "")
    if text in ("-", "—", "None"):
        return Decimal("0")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return Decimal(text)


def parse_percent(value) -> Decimal:
    v = to_decimal(value)
    if v > Decimal("1"):
        v = v / Decimal("100")
    return v


def format_money(value: Decimal) -> str:
    value = q2(value)
    s = f"{value:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def format_pct(value: Decimal) -> str:
    v = (value * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}%"


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).strip().lower())


class PricingEngine:
    def __init__(self, workbook_path: str):
        self.ncm_factors = {}
        self.st_rates = {}
        self.load_workbook(workbook_path)

    def load_workbook(self, workbook_path: str):
        if not os.path.exists(workbook_path):
            raise FileNotFoundError(
                f"Planilha base não encontrada.\n\nColoque o arquivo '{os.path.basename(workbook_path)}' na mesma pasta do programa."
            )

        wb = openpyxl.load_workbook(workbook_path, data_only=False)

        ws_ncm = wb["Base Dados NCM"]
        self.ncm_factors.clear()
        for row in ws_ncm.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.ncm_factors[key] = {
                "fator_custo_filial_35": to_decimal(row[1]),
                "fator_psd_filial_35": to_decimal(row[3]),
                "fator_psd_filial_2_fora_pe": to_decimal(row[4]),
                "fator_psd_filial_4_dentro_rn": to_decimal(row[5]),
                "fator_psd_filial_4_fora_rn": to_decimal(row[6]),
            }

        ws_st = wb["Base ST RN"]
        self.st_rates.clear()
        for row in ws_st.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.st_rates[key] = to_decimal(row[1])

    def calcular(self, compra_para: str, preco_unitario: Decimal, ncm: str, icms: Decimal, ipi: Decimal, frete_tipo: str):
        ncm_key = str(ncm).strip()
        if ncm_key not in self.ncm_factors:
            raise KeyError(f"NCM {ncm_key} não encontrado na aba 'Base Dados NCM'.")

        fatores = self.ncm_factors[ncm_key]
        st_rate = self.st_rates.get(ncm_key, Decimal("0"))

        vlr_icms = preco_unitario * icms
        vlr_ipi = preco_unitario * ipi
        bc_pis_cofins = preco_unitario * (Decimal("1") - icms)
        pis_cofins = (bc_pis_cofins + vlr_ipi) * PIS_COFINS_RATE

        if compra_para == "RN":
            frete_fob = preco_unitario * FRETE_FOB_RATE_RN if frete_tipo == "FOB" else Decimal("0")
            vlr_st = preco_unitario * st_rate
            custo_filial_4 = (preco_unitario + vlr_ipi + vlr_st + frete_fob) - pis_cofins
            custo_filial_2 = custo_filial_4 * Decimal("0.88")
            custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
            r = custo_filial_4 * fatores["fator_psd_filial_4_dentro_rn"]
            s = custo_filial_4 * fatores["fator_psd_filial_4_fora_rn"]
            t = custo_filial_35 * fatores["fator_psd_filial_35"]
            return {"r": r, "s": s, "t": t}

        frete_fob = (preco_unitario + vlr_ipi) * FRETE_FOB_RATE_PE if frete_tipo == "FOB" else Decimal("0")
        custo_filial_2 = (preco_unitario + vlr_ipi + frete_fob) - pis_cofins - vlr_icms
        custo_filial_4 = custo_filial_2 * Decimal("1.12")
        custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
        r = custo_filial_4
        s = custo_filial_2 * fatores["fator_psd_filial_2_fora_pe"]
        t = custo_filial_35 * fatores["fator_psd_filial_35"]
        return {"r": r, "s": s, "t": t}


class UniversalBudgetReader:
    HEADER_ALIASES = {
        "descricao": ["descricao", "descrição", "produto", "item", "nomeproduto", "descricaoproduto", "material"],
        "ncm": ["ncm", "ncmsh", "codigoncm", "classfiscal", "classificacaofiscal", "classificaçãofiscal"],
        "ipi": ["ipi", "aliquotaipi", "percipi", "percentualipi"],
        "icms": ["icms", "aliquotaicms", "percicms", "percentualicms"],
        "frete": ["frete", "tipofrete", "modofrete", "fobcif", "ciffob"],
        "preco": ["preco", "preço", "precounitario", "preçounitário", "valorunitario", "vlunitario", "valor", "unitario", "unitário", "preco rsun"],
        "codigo": ["codigo", "código", "cod", "referencia", "referência"],
        "qtde": ["qtde", "quantidade", "qtd"],
    }

    def read(self, filepath: str):
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            return self._read_excel(filepath)
        if ext == ".csv":
            return self._read_csv(filepath)
        if ext == ".pdf":
            return self._read_pdf(filepath)
        if ext in (".png", ".jpg", ".jpeg"):
            return self._read_image(filepath)
        raise ValueError("Formato não suportado. Use PDF, XLSX, XLSM, CSV, PNG, JPG ou JPEG.")

    def _detect_cnpj_and_mode(self, text_dump: str):
        texto_numeros = re.sub(r"\D", "", text_dump)
        cnpjs = re.findall(r"\d{14}", texto_numeros)
        for cnpj in cnpjs:
            cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            if cnpj == CNPJ_RN_DIGITS or cnpj.endswith("000405"):
                return cnpj_formatado, "RN"
            if cnpj == CNPJ_PE_DIGITS or cnpj.endswith("000154"):
                return cnpj_formatado, "PE"
        upper = text_dump.upper()
        if "KGMLAN" in upper or "KGM LAN" in upper:
            return CNPJ_RN, "RN"
        return "", ""

    def _detect_frete(self, text_dump: str):
        upper = text_dump.upper()
        if "FRETE FOB" in upper or " FOB " in f" {upper} ":
            return "FOB"
        if "FRETE CIF" in upper or " CIF " in f" {upper} ":
            return "CIF"
        return "CIF"

    def _match_header(self, value):
        normalized = normalize_text(value)
        for field, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                if normalize_text(alias) == normalized:
                    return field
        return None

    def _dedupe_items(self, items):
        deduped = []
        seen = set()
        for item in items:
            key = (
                str(item.get("codigo", "")).strip(),
                str(item.get("descricao", "")).strip().upper(),
                re.sub(r"\D", "", str(item.get("ncm", ""))),
                str(q2(to_decimal(item.get("preco", "")))),
                str(item.get("qtde", "")),
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        return deduped

    def _post_process_items(self, items, default_frete):
        processed = []
        for item in self._dedupe_items(items):
            desc = str(item.get("descricao", "")).strip()
            ncm = re.sub(r"\D", "", str(item.get("ncm", "")))
            preco = to_decimal(item.get("preco", ""))
            qtde_dec = to_decimal(item.get("qtde", "1"))
            qtde = int(qtde_dec) if qtde_dec > 0 else 1
            if qtde <= 0 or preco <= 0:
                continue
            if not desc and not ncm:
                continue
            processed.append({
                "codigo": str(item.get("codigo", "")).strip(),
                "descricao": desc,
                "ncm": ncm,
                "ipi": parse_percent(item.get("ipi", "")),
                "icms": parse_percent(item.get("icms", "")),
                "frete": self._detect_frete(str(item.get("frete", ""))) or default_frete,
                "preco": preco,
                "qtde": qtde,
            })
        return processed

    def _extract_from_tabular_rows(self, rows, text_dump):
        found_items = []
        for idx, row in enumerate(rows[:30]):
            current_map = {}
            for col_idx, val in enumerate(row):
                if val in (None, ""):
                    continue
                matched = self._match_header(val)
                if matched:
                    current_map[matched] = col_idx
            if ("descricao" in current_map and "ncm" in current_map and "preco" in current_map and "qtde" in current_map):
                for data_row in rows[idx + 1:]:
                    if not any(v not in (None, "") for v in data_row):
                        continue
                    item = {}
                    for field, col_idx in current_map.items():
                        if col_idx < len(data_row):
                            item[field] = data_row[col_idx]
                    found_items.append(item)
                break
        if found_items:
            return found_items

        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        generic_items = []
        i = 0
        while i < len(lines):
            line = lines[i]
            m_head = re.match(r'^([A-Z0-9\-]+)\s+([\d\.,]+)\s+(\d+)\s+([\d\.,]+)(?:\s+[\d\.]+)?$', line)
            if m_head:
                codigo, preco, qtde, _total = m_head.groups()
                desc_lines = []
                j = i + 1
                while j < len(lines) and not re.search(r'\b\d{8}\b', lines[j]):
                    desc_lines.append(lines[j])
                    j += 1
                if j < len(lines):
                    tail = lines[j]
                    m_tail = re.search(r'(\d{8}).*?(\d{1,2}(?:[.,]\d{1,2})?).*?(\d{1,2}(?:[.,]\d{1,2})?)', tail)
                    if m_tail:
                        ncm, icms, ipi = m_tail.groups()
                        generic_items.append({
                            "codigo": codigo,
                            "descricao": " ".join(desc_lines).replace("#", "").strip(),
                            "ncm": ncm,
                            "icms": icms,
                            "ipi": ipi,
                            "preco": preco,
                            "frete": self._detect_frete(text_dump),
                            "qtde": qtde,
                        })
                        i = j
            i += 1
        return generic_items

    def _read_excel(self, filepath: str):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        text_parts, found_items = [], []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            sheet_text = []
            for row in rows:
                row_text = " ".join([str(v) for v in row if v not in (None, "")])
                if row_text:
                    sheet_text.append(row_text)
                    text_parts.append(row_text)
            found_items.extend(self._extract_from_tabular_rows(rows, "\n".join(sheet_text)))
        text_dump = "\n".join(text_parts)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}

    def _read_csv(self, filepath: str):
        with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        text_dump = "\n".join([" ".join([str(v) for v in row if v not in (None, "")]) for row in rows])
        found_items = self._extract_from_tabular_rows(rows, text_dump)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}

    def _read_pdf(self, filepath: str):
        try:
            import pdfplumber
        except Exception:
            raise ImportError("Para ler PDF, instale pdfplumber: python -m pip install pdfplumber")
        text_parts, found_items = [], []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                text_parts.append(text)
                tables = page.extract_tables() or []
                for table in tables:
                    if table:
                        found_items.extend(self._extract_from_tabular_rows(table, text))
        text_dump = "\n".join(text_parts)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}

    def _read_image(self, filepath: str):
        try:
            from PIL import Image
            import pytesseract
        except Exception:
            raise ImportError("Para ler imagem, instale Pillow e pytesseract, além do Tesseract OCR no Windows.")
        image = Image.open(filepath)
        text_dump = pytesseract.image_to_string(image, lang="por+eng", config="--psm 6")
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        found_items = []
        current = {"codigo": "", "descricao": "", "ncm": "", "icms": "", "ipi": "", "preco": "", "frete": frete, "qtde": "1"}
        for line in lines:
            upper = line.upper()
            if not current["descricao"] and not any(x in upper for x in ["NCM", "ICMS", "IPI", "CNPJ", "R$"]):
                current["descricao"] = line
            if not current["ncm"]:
                m = re.search(r"\b(\d{8})\b", line)
                if m:
                    current["ncm"] = m.group(1)
            if not current["icms"] and "ICMS" in upper:
                p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                if p:
                    current["icms"] = p.group(1)
            if not current["ipi"] and "IPI" in upper:
                p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                if p:
                    current["ipi"] = p.group(1)
            if not current["preco"] and ("R$" in upper or "UNIT" in upper or "VALOR" in upper or "PRECO" in upper or "PREÇO" in upper):
                m = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})', line)
                if m:
                    current["preco"] = m.group(1)
            if current["qtde"] == "1":
                m = re.search(r'\bQTD(?:E)?\s*:?\s*(\d+)\b', upper)
                if m:
                    current["qtde"] = m.group(1)
        if current["descricao"] or current["ncm"]:
            found_items.append(current)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1360x860")
        self.minsize(1220, 760)

        self.engine = PricingEngine(DEFAULT_WORKBOOK)
        self.reader = UniversalBudgetReader()
        self.items = []
        self.current_results = []

        self._build_styles()
        self._build_ui()

    def _build_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except Exception:
            pass
        style.configure("Header.TLabel", background="#92278f", foreground="white", font=("Segoe UI", 20, "bold"), padding=14)
        style.configure("Section.TLabelframe.Label", font=("Segoe UI", 11, "bold"))

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        ttk.Label(self, text="SIMULADOR DE PRECIFICAÇÃO - VERSÃO UNIVERSAL", style="Header.TLabel", anchor="center").grid(row=0, column=0, sticky="ew")

        main = ttk.Frame(self, padding=12)
        main.grid(row=1, column=0, sticky="nsew")
        main.columnconfigure(0, weight=1)
        main.rowconfigure(2, weight=1)

        top = ttk.LabelFrame(main, text="Arquivo do orçamento", style="Section.TLabelframe", padding=12)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        self.budget_path_var = tk.StringVar()
        ttk.Button(top, text="Selecionar arquivo", command=self.open_budget).grid(row=0, column=0, padx=(0, 8), sticky="w")
        ttk.Entry(top, textvariable=self.budget_path_var).grid(row=0, column=1, sticky="ew")

        info = ttk.Frame(top)
        info.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        info.columnconfigure((0, 1, 2, 3), weight=1)

        self.detect_cnpj_var = tk.StringVar(value="CNPJ: -")
        self.detect_mode_var = tk.StringVar(value="Compra para: -")
        self.detect_frete_var = tk.StringVar(value="Frete: -")
        self.total_itens_var = tk.StringVar(value="Itens: 0")

        tk.Label(info, textvariable=self.detect_cnpj_var, anchor="w", font=("Segoe UI", 10, "bold"), bg="#f3f3f3", padx=10, pady=8).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        tk.Label(info, textvariable=self.detect_mode_var, anchor="w", font=("Segoe UI", 10, "bold"), bg="#f3f3f3", padx=10, pady=8).grid(row=0, column=1, sticky="ew", padx=6)
        tk.Label(info, textvariable=self.detect_frete_var, anchor="w", font=("Segoe UI", 10, "bold"), bg="#f3f3f3", padx=10, pady=8).grid(row=0, column=2, sticky="ew", padx=6)
        tk.Label(info, textvariable=self.total_itens_var, anchor="w", font=("Segoe UI", 10, "bold"), bg="#f3f3f3", padx=10, pady=8).grid(row=0, column=3, sticky="ew", padx=(6, 0))

        actions = ttk.Frame(main)
        actions.grid(row=1, column=0, sticky="ew", pady=(10, 10))
        actions.columnconfigure(0, weight=1)

        left = ttk.Frame(actions)
        left.grid(row=0, column=0, sticky="w")
        ttk.Label(left, text="Item selecionado:").pack(side="left")
        self.item_combo_var = tk.StringVar()
        self.item_combo = ttk.Combobox(left, textvariable=self.item_combo_var, state="readonly", width=95)
        self.item_combo.pack(side="left", padx=(8, 0))
        self.item_combo.bind("<<ComboboxSelected>>", self.on_item_selected)

        right = ttk.Frame(actions)
        right.grid(row=0, column=1, sticky="e")
        ttk.Button(right, text="Calcular orçamento inteiro", command=self.calcular_orcamento_inteiro).pack(side="left", padx=(0, 8))
        ttk.Button(right, text="Exportar Excel", command=self.exportar_excel).pack(side="left")

        table_box = ttk.LabelFrame(main, text="Todos os itens precificados", style="Section.TLabelframe", padding=8)
        table_box.grid(row=2, column=0, sticky="nsew")
        table_box.columnconfigure(0, weight=1)
        table_box.rowconfigure(0, weight=1)

        cols = ("idx", "codigo", "produto", "qtde", "ncm", "preco", "icms", "ipi", "frete", "r", "s", "t")
        self.tree = ttk.Treeview(table_box, columns=cols, show="headings", height=24)

        headers = {
            "idx": "#",
            "codigo": "Código",
            "produto": "Produto",
            "qtde": "Qtde",
            "ncm": "NCM",
            "preco": "Preço Unitário",
            "icms": "ICMS",
            "ipi": "IPI",
            "frete": "Frete",
            "r": "R / Filial 4",
            "s": "S / Filial 2",
            "t": "T / Filial 3 e 5",
        }
        widths = {
            "idx": 40,
            "codigo": 90,
            "produto": 720,
            "qtde": 60,
            "ncm": 100,
            "preco": 110,
            "icms": 70,
            "ipi": 70,
            "frete": 70,
            "r": 120,
            "s": 120,
            "t": 130,
        }

        for col in cols:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor="center")

        self.tree.column("produto", anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")

        self.tree.tag_configure("price_highlight", background="#fff4cc")
        self.tree.tag_configure("normal_row", background="#ffffff")

        yscroll = ttk.Scrollbar(table_box, orient="vertical", command=self.tree.yview)
        yscroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.status_var = tk.StringVar(value="Selecione um arquivo para listar todos os itens precificados.")
        status = tk.Label(main, textvariable=self.status_var, anchor="w", font=("Segoe UI", 10, "bold"), bg="#e9e4ee", padx=10, pady=8)
        status.grid(row=3, column=0, sticky="ew", pady=(10, 0))

    def open_budget(self):
        path = filedialog.askopenfilename(
            title="Selecione o arquivo do orçamento",
            filetypes=[
                ("Arquivos suportados", "*.pdf *.xlsx *.xlsm *.csv *.png *.jpg *.jpeg"),
                ("PDF", "*.pdf"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Imagem", "*.png *.jpg *.jpeg"),
            ],
        )
        if not path:
            return
        try:
            data = self.reader.read(path)
            self.budget_path_var.set(path)
            self.detect_cnpj_var.set(f"CNPJ: {data.get('cnpj') or 'não identificado'}")
            self.detect_mode_var.set(f"Compra para: {data.get('compra_para') or 'não identificada'}")
            self.detect_frete_var.set(f"Frete: {data.get('frete') or '-'}")
            self.items = data.get("items", [])

            if not self.items:
                raise ValueError("Não consegui identificar itens automaticamente nesse arquivo.")

            self.total_itens_var.set(f"Itens: {len(self.items)}")

            labels = []
            for i, item in enumerate(self.items, start=1):
                desc = item["descricao"] or "(sem descrição)"
                labels.append(f"{i}. {item.get('codigo','')} | Qtde {item.get('qtde',1)} | {desc[:90]} | NCM: {item['ncm']}")
            self.item_combo["values"] = labels
            if labels:
                self.item_combo.current(0)

            self.status_var.set(f"Arquivo lido com sucesso. Itens únicos encontrados: {len(self.items)}.")
            self.calcular_orcamento_inteiro()

        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def on_item_selected(self, event=None):
        idx = self.item_combo.current()
        if idx >= 0 and self.tree.get_children():
            self.tree.selection_set(self.tree.get_children()[idx])
            self.tree.see(self.tree.get_children()[idx])

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        idx = int(values[0]) - 1
        if 0 <= idx < len(self.items):
            self.item_combo.current(idx)

    def calcular_orcamento_inteiro(self):
        if not self.items:
            messagebox.showwarning(APP_TITLE, "Selecione primeiro um arquivo.")
            return

        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        self.current_results = []
        compra_para = ""
        cnpj_text = self.detect_mode_var.get().upper()
        if "RN" in cnpj_text:
            compra_para = "RN"
        elif "PE" in cnpj_text:
            compra_para = "PE"

        if compra_para not in ("RN", "PE"):
            raise ValueError("Não foi possível identificar automaticamente RN ou PE pelo CNPJ do arquivo.")

        for idx, item in enumerate(self.items, start=1):
            try:
                result = self.engine.calcular(
                    compra_para,
                    item["preco"],
                    item["ncm"],
                    item["icms"],
                    item["ipi"],
                    item["frete"],
                )
                record = {
                    "idx": idx,
                    "codigo": item["codigo"],
                    "produto": item["descricao"],
                    "qtde": item["qtde"],
                    "ncm": item["ncm"],
                    "preco": item["preco"],
                    "icms": item["icms"],
                    "ipi": item["ipi"],
                    "frete": item["frete"],
                    "r": result["r"],
                    "s": result["s"],
                    "t": result["t"],
                }
                self.current_results.append(record)
                self.tree.insert(
                    "",
                    "end",
                    values=(
                        record["idx"],
                        record["codigo"],
                        record["produto"],
                        record["qtde"],
                        record["ncm"],
                        format_money(record["preco"]),
                        format_pct(record["icms"]),
                        format_pct(record["ipi"]),
                        record["frete"],
                        format_money(record["r"]),
                        format_money(record["s"]),
                        format_money(record["t"]),
                    ),
                    tags=("price_highlight",)
                )
            except Exception as e:
                self.tree.insert(
                    "",
                    "end",
                    values=(idx, item["codigo"], f"ERRO: {str(e)}", item.get("qtde", 1), item["ncm"], "", "", "", item["frete"], "", "", ""),
                    tags=("normal_row",)
                )

        self.status_var.set(f"Todos os itens foram precificados. Total de itens únicos: {len(self.items)}.")

    def exportar_excel(self):
        if not self.current_results:
            messagebox.showwarning(APP_TITLE, "Calcule o orçamento inteiro antes de exportar.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Salvar resultado em Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="resultado_precificacao_universal_layout.xlsx",
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultado"

        headers = ["Item", "Codigo", "Produto", "Qtde", "NCM", "Preco Unitario", "ICMS", "IPI", "Frete", "Resultado R", "Resultado S", "Resultado T"]
        ws.append(headers)

        for rec in self.current_results:
            ws.append([
                rec["idx"],
                rec["codigo"],
                rec["produto"],
                rec["qtde"],
                rec["ncm"],
                float(q2(rec["preco"])),
                float(rec["icms"]),
                float(rec["ipi"]),
                rec["frete"],
                float(q2(rec["r"])),
                float(q2(rec["s"])),
                float(q2(rec["t"])),
            ])

        header_fill = PatternFill(fill_type="solid", fgColor="92278F")
        header_font = Font(color="FFFFFF", bold=True)
        price_fill = PatternFill(fill_type="solid", fgColor="FFF4CC")
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        money_cols = {"F", "J", "K", "L"}
        pct_cols = {"G", "H"}

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column_letter in money_cols and cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'
                elif cell.column_letter in pct_cols and cell.value is not None:
                    cell.number_format = '0.00%'
                if cell.column_letter == "F":
                    cell.fill = price_fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 65)

        ws.freeze_panes = "A2"
        wb.save(filepath)
        messagebox.showinfo(APP_TITLE, f"Arquivo exportado com sucesso:\n{filepath}")


if __name__ == "__main__":
    app = App()
    app.mainloop()