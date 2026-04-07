import os
import re
import csv
from decimal import Decimal, ROUND_HALF_UP
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_TITLE = "Assistente Profissional de Orcamento"
DEFAULT_WORKBOOK = os.path.join(os.path.dirname(os.path.abspath(__file__)), "1 NOVO SIMULADOR_PRECIFICACAO_V2.xlsx")

PIS_COFINS_RATE = Decimal("0.0925")
FRETE_FOB_RATE_RN = Decimal("0.10")
FRETE_FOB_RATE_PE = Decimal("0.05")
ROUND_MONEY = Decimal("0.01")

CNPJ_RN = "18.217.682/0004-05"
CNPJ_PE = "18.217.682/0001-54"
CNPJ_RN_DIGITS = "18217682000405"
CNPJ_PE_DIGITS = "18217682000154"


def q2(value: Decimal) -> Decimal:
    return value.quantize(ROUND_MONEY, rounding=ROUND_HALF_UP)


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip().replace("R$", "").replace("%", "").replace(" ", "")
    if text in ("-", "—", "None"):
        return Decimal("0")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return Decimal(text)


def parse_percent(value) -> Decimal:
    v = to_decimal(value)
    if v > Decimal("1"):
        v = v / Decimal("100")
    return v


def format_money(value: Decimal) -> str:
    value = q2(value)
    s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def format_pct(value: Decimal) -> str:
    v = (value * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}%"


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).strip().lower())


class PricingEngine:
    def __init__(self, workbook_path: str):
        self.ncm_factors = {}
        self.st_rates = {}
        self.load_workbook(workbook_path)

    def load_workbook(self, workbook_path: str):
        if not os.path.exists(workbook_path):
            raise FileNotFoundError(
                f"Planilha base nao encontrada.\n\nColoque o arquivo '{os.path.basename(workbook_path)}' na mesma pasta do programa."
            )

        wb = openpyxl.load_workbook(workbook_path, data_only=False)

        ws_ncm = wb["Base Dados NCM"]
        self.ncm_factors.clear()
        for row in ws_ncm.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.ncm_factors[key] = {
                "fator_custo_filial_35": to_decimal(row[1]),
                "fator_psd_filial_35": to_decimal(row[3]),
                "fator_psd_filial_2_fora_pe": to_decimal(row[4]),
                "fator_psd_filial_4_dentro_rn": to_decimal(row[5]),
                "fator_psd_filial_4_fora_rn": to_decimal(row[6]),
            }

        ws_st = wb["Base ST RN"]
        self.st_rates.clear()
        for row in ws_st.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.st_rates[key] = to_decimal(row[1])

    def calcular(self, compra_para: str, preco_unitario: Decimal, ncm: str, icms: Decimal, ipi: Decimal, frete_tipo: str):
        ncm_key = str(ncm).strip()
        if ncm_key not in self.ncm_factors:
            raise KeyError(f"NCM {ncm_key} nao encontrado na aba 'Base Dados NCM'.")

        fatores = self.ncm_factors[ncm_key]
        st_rate = self.st_rates.get(ncm_key, Decimal("0"))

        vlr_icms = preco_unitario * icms
        vlr_ipi = preco_unitario * ipi
        bc_pis_cofins = preco_unitario * (Decimal("1") - icms)
        pis_cofins = (bc_pis_cofins + vlr_ipi) * PIS_COFINS_RATE

        if compra_para == "RN":
            frete_fob = preco_unitario * FRETE_FOB_RATE_RN if frete_tipo == "FOB" else Decimal("0")
            vlr_st = preco_unitario * st_rate
            custo_filial_4 = (preco_unitario + vlr_ipi + vlr_st + frete_fob) - pis_cofins
            custo_filial_2 = custo_filial_4 * Decimal("0.88")
            custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
            r = custo_filial_4 * fatores["fator_psd_filial_4_dentro_rn"]
            s = custo_filial_4 * fatores["fator_psd_filial_4_fora_rn"]
            t = custo_filial_35 * fatores["fator_psd_filial_35"]
            return {"r": r, "s": s, "t": t}

        frete_fob = (preco_unitario + vlr_ipi) * FRETE_FOB_RATE_PE if frete_tipo == "FOB" else Decimal("0")
        custo_filial_2 = (preco_unitario + vlr_ipi + frete_fob) - pis_cofins - vlr_icms
        custo_filial_4 = custo_filial_2 * Decimal("1.12")
        custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
        r = custo_filial_4
        s = custo_filial_2 * fatores["fator_psd_filial_2_fora_pe"]
        t = custo_filial_35 * fatores["fator_psd_filial_35"]
        return {"r": r, "s": s, "t": t}


class UniversalBudgetReader:
    HEADER_ALIASES = {
        "descricao": ["descricao", "descrição", "produto", "item", "nomeproduto", "descricaoproduto", "material"],
        "ncm": ["ncm", "ncmsh", "codigoncm", "classfiscal", "classificacaofiscal", "classificaçãofiscal"],
        "ipi": ["ipi", "aliquotaipi", "percipi", "percentualipi"],
        "icms": ["icms", "aliquotaicms", "percicms", "percentualicms"],
        "frete": ["frete", "tipofrete", "modofrete", "fobcif", "ciffob"],
        "preco": ["preco", "preço", "precounitario", "preçounitário", "valorunitario", "vlunitario", "valor", "unitario", "unitário", "preco rsun"],
        "codigo": ["codigo", "código", "cod", "referencia", "referência"],
        "qtde": ["qtde", "quantidade", "qtd"],
    }

    def read(self, filepath: str):
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            return self._read_excel(filepath)
        if ext == ".csv":
            return self._read_csv(filepath)
        if ext == ".pdf":
            return self._read_pdf(filepath)
        if ext in (".png", ".jpg", ".jpeg"):
            return self._read_image(filepath)
        raise ValueError("Formato nao suportado. Use PDF, XLSX, XLSM, CSV, PNG, JPG ou JPEG.")

    def _detect_cnpj_and_mode(self, text_dump: str):
        texto_numeros = re.sub(r"\D", "", text_dump)
        cnpjs = re.findall(r"\d{14}", texto_numeros)
        for cnpj in cnpjs:
            cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            if cnpj == CNPJ_RN_DIGITS or cnpj.endswith("000405"):
                return cnpj_formatado, "RN"
            if cnpj == CNPJ_PE_DIGITS or cnpj.endswith("000154"):
                return cnpj_formatado, "PE"
        upper = text_dump.upper()
        if "KGMLAN" in upper or "KGM LAN" in upper:
            return CNPJ_RN, "RN"
        return "", ""

    def _detect_frete(self, text_dump: str):
        upper = text_dump.upper()
        if "FRETE FOB" in upper or " FOB " in f" {upper} ":
            return "FOB"
        if "FRETE CIF" in upper or " CIF " in f" {upper} ":
            return "CIF"
        return "CIF"

    def _match_header(self, value):
        normalized = normalize_text(value)
        for field, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                if normalize_text(alias) == normalized:
                    return field
        return None

    def _dedupe_items(self, items):
        deduped = []
        seen = set()
        for item in items:
            key = (
                str(item.get("codigo", "")).strip(),
                str(item.get("descricao", "")).strip().upper(),
                re.sub(r"\D", "", str(item.get("ncm", ""))),
                str(q2(to_decimal(item.get("preco", "")))),
                str(item.get("qtde", "")),
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        return deduped

    def _post_process_items(self, items, default_frete):
        processed = []
        for item in self._dedupe_items(items):
            desc = str(item.get("descricao", "")).strip()
            ncm = re.sub(r"\D", "", str(item.get("ncm", "")))
            preco = to_decimal(item.get("preco", ""))
            qtde_dec = to_decimal(item.get("qtde", "1"))
            qtde = int(qtde_dec) if qtde_dec > 0 else 1
            if qtde <= 0 or preco <= 0:
                continue
            if not desc and not ncm:
                continue
            processed.append({
                "codigo": str(item.get("codigo", "")).strip(),
                "descricao": desc,
                "ncm": ncm,
                "ipi": parse_percent(item.get("ipi", "")),
                "icms": parse_percent(item.get("icms", "")),
                "frete": self._detect_frete(str(item.get("frete", ""))) or default_frete,
                "preco": preco,
                "qtde": qtde,
            })
        return processed

    def _extract_from_tabular_rows(self, rows, text_dump):
        found_items = []
        for idx, row in enumerate(rows[:30]):
            current_map = {}
            for col_idx, val in enumerate(row):
                if val in (None, ""):
                    continue
                matched = self._match_header(val)
                if matched:
                    current_map[matched] = col_idx
            if ("descricao" in current_map and "ncm" in current_map and "preco" in current_map and "qtde" in current_map):
                for data_row in rows[idx + 1:]:
                    if not any(v not in (None, "") for v in data_row):
                        continue
                    item = {}
                    for field, col_idx in current_map.items():
                        if col_idx < len(data_row):
                            item[field] = data_row[col_idx]
                    found_items.append(item)
                break
        if found_items:
            return found_items

        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        generic_items = []
        i = 0
        while i < len(lines):
            line = lines[i]
            m_head = re.match(r'^([A-Z0-9\-]+)\s+([\d\.,]+)\s+(\d+)\s+([\d\.,]+)(?:\s+[\d\.]+)?$', line)
            if m_head:
                codigo, preco, qtde, _total = m_head.groups()
                desc_lines = []
                j = i + 1
                while j < len(lines) and not re.search(r'\b\d{8}\b', lines[j]):
                    desc_lines.append(lines[j])
                    j += 1
                if j < len(lines):
                    tail = lines[j]
                    m_tail = re.search(r'(\d{8}).*?(\d{1,2}(?:[.,]\d{1,2})?).*?(\d{1,2}(?:[.,]\d{1,2})?)', tail)
                    if m_tail:
                        ncm, icms, ipi = m_tail.groups()
                        generic_items.append({
                            "codigo": codigo,
                            "descricao": " ".join(desc_lines).replace("#", "").strip(),
                            "ncm": ncm,
                            "icms": icms,
                            "ipi": ipi,
                            "preco": preco,
                            "frete": self._detect_frete(text_dump),
                            "qtde": qtde,
                        })
                        i = j
            i += 1
        return generic_items

    def _read_excel(self, filepath: str):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        text_parts, found_items = [], []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            sheet_text = []
            for row in rows:
                row_text = " ".join([str(v) for v in row if v not in (None, "")])
                if row_text:
                    sheet_text.append(row_text)
                    text_parts.append(row_text)
            found_items.extend(self._extract_from_tabular_rows(rows, "\n".join(sheet_text)))
        text_dump = "\n".join(text_parts)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}

    def _read_csv(self, filepath: str):
        with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        text_dump = "\n".join([" ".join([str(v) for v in row if v not in (None, "")]) for row in rows])
        found_items = self._extract_from_tabular_rows(rows, text_dump)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}

    def _read_pdf(self, filepath: str):
        try:
            import pdfplumber
        except Exception:
            raise ImportError("Para ler PDF, instale pdfplumber: python -m pip install pdfplumber")
        text_parts, found_items = [], []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                text_parts.append(text)
                tables = page.extract_tables() or []
                for table in tables:
                    if table:
                        found_items.extend(self._extract_from_tabular_rows(table, text))
        text_dump = "\n".join(text_parts)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}

    def _read_image(self, filepath: str):
        try:
            from PIL import Image
            import pytesseract
        except Exception:
            raise ImportError("Para ler imagem, instale Pillow e pytesseract, alem do Tesseract OCR no Windows.")
        image = Image.open(filepath)
        text_dump = pytesseract.image_to_string(image, lang="por+eng", config="--psm 6")
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        found_items = []
        current = {"codigo": "", "descricao": "", "ncm": "", "icms": "", "ipi": "", "preco": "", "frete": frete, "qtde": "1"}
        for line in lines:
            upper = line.upper()
            if not current["descricao"] and not any(x in upper for x in ["NCM", "ICMS", "IPI", "CNPJ", "R$"]):
                current["descricao"] = line
            if not current["ncm"]:
                m = re.search(r"\b(\d{8})\b", line)
                if m:
                    current["ncm"] = m.group(1)
            if not current["icms"] and "ICMS" in upper:
                p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                if p:
                    current["icms"] = p.group(1)
            if not current["ipi"] and "IPI" in upper:
                p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                if p:
                    current["ipi"] = p.group(1)
            if not current["preco"] and ("R$" in upper or "UNIT" in upper or "VALOR" in upper or "PRECO" in upper or "PREÇO" in upper):
                m = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})', line)
                if m:
                    current["preco"] = m.group(1)
            if current["qtde"] == "1":
                m = re.search(r'\bQTD(?:E)?\s*:?\s*(\d+)\b', upper)
                if m:
                    current["qtde"] = m.group(1)
        if current["descricao"] or current["ncm"]:
            found_items.append(current)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete)}


class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = tk.Canvas(self, highlightthickness=0, bg="#f4f6fb")
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vsb.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.window, width=event.width)

    def _on_mousewheel(self, event):
        if self.winfo_ismapped():
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1480x900")
        self.minsize(1320, 820)
        self.configure(bg="#eceff5")

        self.engine = PricingEngine(DEFAULT_WORKBOOK)
        self.reader = UniversalBudgetReader()
        self.items = []
        self.current_results = []

        self._build_styles()
        self._build_layout()
        self.show_home()

    def _build_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except Exception:
            pass

        style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=(14, 10))
        style.configure("Ghost.TButton", font=("Segoe UI", 10), padding=(12, 10))
        style.configure("Card.TLabelframe", background="#ffffff")
        style.configure("Card.TLabelframe.Label", font=("Segoe UI", 11, "bold"))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30)

    def _build_layout(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        header = tk.Frame(self, bg="#6f2dbd", height=110)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)
        tk.Label(header, text="ASSISTENTE PROFISSIONAL DE ORCAMENTO", bg="#6f2dbd", fg="white", font=("Segoe UI", 24, "bold")).pack(pady=(20, 4))
        tk.Label(header, text="Entrada manual ou leitura automatica de orcamento com visual corporativo", bg="#6f2dbd", fg="#eadcff", font=("Segoe UI", 11)).pack()

        nav = tk.Frame(self, bg="#eceff5", pady=12)
        nav.grid(row=1, column=0, sticky="ew")
        nav.grid_columnconfigure(4, weight=1)
        ttk.Button(nav, text="Tela inicial", style="Ghost.TButton", command=self.show_home).grid(row=0, column=0, padx=(20, 10))
        ttk.Button(nav, text="Digitar manualmente", style="Ghost.TButton", command=self.show_manual).grid(row=0, column=1, padx=10)
        ttk.Button(nav, text="Anexar orcamento", style="Ghost.TButton", command=self.show_attach).grid(row=0, column=2, padx=10)
        ttk.Button(nav, text="Exportar Excel", style="Primary.TButton", command=self.exportar_excel).grid(row=0, column=3, padx=10)

        self.content = tk.Frame(self, bg="#eceff5")
        self.content.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 10))
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(0, weight=1)

        self.status_var = tk.StringVar(value="Pronto para iniciar.")
        status = tk.Label(self, textvariable=self.status_var, anchor="w", bg="#dfe5f2", fg="#16253d", font=("Segoe UI", 10, "bold"), padx=14, pady=10)
        status.grid(row=3, column=0, sticky="ew")

    def clear_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    def build_hero(self, parent, title, subtitle):
        hero = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        hero.grid(row=0, column=0, sticky="ew")
        hero.grid_columnconfigure(0, weight=1)
        tk.Label(hero, text=title, bg="#ffffff", fg="#10243e", font=("Segoe UI", 22, "bold")).grid(row=0, column=0, sticky="w", padx=24, pady=(22, 6))
        tk.Label(hero, text=subtitle, bg="#ffffff", fg="#64748b", font=("Segoe UI", 11)).grid(row=1, column=0, sticky="w", padx=24, pady=(0, 22))
        return hero

    def show_home(self):
        self.clear_content()
        self.content.grid_rowconfigure(1, weight=1)
        self.build_hero(self.content, "Escolha como deseja iniciar", "Use o modo manual para informar um item rapidamente ou anexe um orcamento para leitura automatica.")

        cards = tk.Frame(self.content, bg="#eceff5")
        cards.grid(row=1, column=0, sticky="nsew", pady=(16, 0))
        cards.grid_columnconfigure((0, 1), weight=1)

        self._create_option_card(cards, 0, "Modo manual", "Preencha Nome do Produto, NCM, IPI, ICMS, Preco, Quantidade, Frete e Filial. Ideal para cotacao rapida de um item.", "Abrir formulario", self.show_manual)
        self._create_option_card(cards, 1, "Anexar orcamento", "Selecione arquivos PDF, XLSX, XLSM, CSV, PNG, JPG ou JPEG. O sistema tenta identificar CNPJ, frete, filial e todos os itens automaticamente.", "Selecionar arquivo", self.show_attach)
        self.status_var.set("Tela inicial carregada.")

    def _create_option_card(self, parent, col, title, desc, btn_text, command):
        card = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 8, 8 if col == 0 else 0))
        card.grid_columnconfigure(0, weight=1)
        tk.Label(card, text=title, bg="#ffffff", fg="#10243e", font=("Segoe UI", 18, "bold")).grid(row=0, column=0, sticky="w", padx=24, pady=(24, 10))
        tk.Label(card, text=desc, bg="#ffffff", fg="#475569", font=("Segoe UI", 11), justify="left", wraplength=520).grid(row=1, column=0, sticky="w", padx=24)
        ttk.Button(card, text=btn_text, style="Primary.TButton", command=command).grid(row=2, column=0, sticky="w", padx=24, pady=24)

    def show_manual(self):
        self.clear_content()
        wrapper = tk.Frame(self.content, bg="#eceff5")
        wrapper.grid(row=0, column=0, sticky="nsew")
        wrapper.grid_columnconfigure(0, weight=1)
        wrapper.grid_rowconfigure(2, weight=1)
        self.build_hero(wrapper, "Entrada manual", "Preencha os campos abaixo. O formulario e os botoes principais ficam sempre visiveis na parte superior.")

        body = tk.Frame(wrapper, bg="#eceff5")
        body.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)

        left_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left_card.grid_columnconfigure(0, weight=1)
        tk.Label(left_card, text="Formulario", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=(18, 10))

        form = tk.Frame(left_card, bg="#ffffff")
        form.grid(row=1, column=0, sticky="ew", padx=22)
        form.grid_columnconfigure(1, weight=1)

        self.manual_nome = tk.StringVar()
        self.manual_ncm = tk.StringVar()
        self.manual_ipi = tk.StringVar(value="0")
        self.manual_icms = tk.StringVar(value="7")
        self.manual_preco = tk.StringVar()
        self.manual_qtde = tk.StringVar(value="1")
        self.manual_frete = tk.StringVar(value="FOB")
        self.manual_filial = tk.StringVar(value=f"Natal - {CNPJ_RN}")

        self._add_entry(form, "Nome do Produto", self.manual_nome, 0, 0, 1)
        self._add_entry(form, "NCM", self.manual_ncm, 1, 0, 1)
        self._add_entry(form, "IPI (%)", self.manual_ipi, 2, 0, 1)
        self._add_entry(form, "ICMS (%)", self.manual_icms, 3, 0, 1)
        self._add_entry(form, "Preco Unitario", self.manual_preco, 4, 0, 1)
        self._add_entry(form, "Quantidade", self.manual_qtde, 5, 0, 1)

        tk.Label(form, text="Frete", bg="#ffffff", fg="#16253d", font=("Segoe UI", 11, "bold")).grid(row=6, column=0, sticky="w", pady=(10, 4))
        frete_box = tk.Frame(form, bg="#ffffff")
        frete_box.grid(row=6, column=1, sticky="w", pady=(10, 4), padx=(12, 0))
        ttk.Radiobutton(frete_box, text="FOB", variable=self.manual_frete, value="FOB").pack(side="left", padx=(0, 16))
        ttk.Radiobutton(frete_box, text="CIF", variable=self.manual_frete, value="CIF").pack(side="left")

        tk.Label(form, text="Filial", bg="#ffffff", fg="#16253d", font=("Segoe UI", 11, "bold")).grid(row=7, column=0, sticky="w", pady=(12, 4))
        filial_cb = ttk.Combobox(form, textvariable=self.manual_filial, state="readonly", width=54)
        filial_cb["values"] = [f"Natal - {CNPJ_RN}", f"Pernambuco - {CNPJ_PE}"]
        filial_cb.grid(row=7, column=1, sticky="ew", pady=(12, 4), padx=(12, 0))

        tk.Label(left_card, text="Dica: preencha NCM, IPI, ICMS, Preco Unitario e Filial. Depois clique em Confirmar item manual.", bg="#fff4cc", fg="#5b4a00", font=("Segoe UI", 10), wraplength=720, padx=14, pady=12).grid(row=2, column=0, sticky="ew", padx=22, pady=(16, 10))

        fixed_actions = tk.Frame(left_card, bg="#ffffff")
        fixed_actions.grid(row=3, column=0, sticky="ew", padx=22, pady=(0, 18))
        ttk.Button(fixed_actions, text="Limpar campos", style="Ghost.TButton", command=self.limpar_manual).pack(side="left")
        ttk.Button(fixed_actions, text="Confirmar item manual", style="Primary.TButton", command=self.confirmar_manual).pack(side="right")

        right_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        right_card.grid(row=0, column=1, sticky="nsew")
        right_card.grid_columnconfigure(0, weight=1)
        tk.Label(right_card, text="Resumo do modo manual", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=(18, 10))
        steps_text = (
            "1. Digite as informacoes do produto.

"
            "2. Escolha se o frete e FOB ou CIF.

"
            f"3. Selecione a filial: Natal ({CNPJ_RN}) ou Pernambuco ({CNPJ_PE}).

"
            "4. Clique em Confirmar item manual para gerar o item.

"
            "5. O item sera enviado para a tabela de resultados abaixo.

"
            "6. Exporte para Excel quando desejar."
        )
        tk.Label(right_card, text=steps_text, justify="left", bg="#eef2ff", fg="#334155", font=("Segoe UI", 11), wraplength=420, padx=18, pady=18).grid(row=1, column=0, sticky="ew", padx=22)
        tk.Label(right_card, text="Nesta versao, o formulario fica sempre visivel. A grade de resultados permanece abaixo da area de entrada.", justify="left", bg="#fff7e6", fg="#7c5a00", font=("Segoe UI", 10), wraplength=420, padx=16, pady=14).grid(row=2, column=0, sticky="ew", padx=22, pady=(18, 18))

        self._build_results_area(wrapper, row=2)
        self.status_var.set("Modo manual ativo. Preencha os campos e clique em Confirmar item manual.")

    def _add_entry(self, parent, label, variable, row, col, span=1):
        tk.Label(parent, text=label, bg="#ffffff", fg="#16253d", font=("Segoe UI", 11, "bold")).grid(row=row, column=col, sticky="w", pady=(10, 4))
        entry = ttk.Entry(parent, textvariable=variable, font=("Segoe UI", 11))
        entry.grid(row=row, column=col + 1, sticky="ew", pady=(10, 4), padx=(12, 0), columnspan=span)
        return entry

    def show_attach(self):
        self.clear_content()
        wrapper = tk.Frame(self.content, bg="#eceff5")
        wrapper.grid(row=0, column=0, sticky="nsew")
        wrapper.grid_columnconfigure(0, weight=1)
        wrapper.grid_rowconfigure(2, weight=1)
        self.build_hero(wrapper, "Leitura automatica de orcamento", "Anexe um arquivo e o sistema vai tentar identificar CNPJ, frete, filial e itens automaticamente.")

        body = tk.Frame(wrapper, bg="#eceff5")
        body.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        body.grid_columnconfigure(0, weight=2)
        body.grid_columnconfigure(1, weight=3)

        left_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left_card.grid_columnconfigure(1, weight=1)

        tk.Label(left_card, text="Anexar orcamento", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=22, pady=(18, 10))
        self.budget_path_var = tk.StringVar()
        ttk.Button(left_card, text="Selecionar arquivo", style="Primary.TButton", command=self.open_budget).grid(row=1, column=0, padx=(22, 12), pady=(0, 12), sticky="w")
        ttk.Entry(left_card, textvariable=self.budget_path_var, font=("Segoe UI", 10)).grid(row=1, column=1, padx=(0, 22), pady=(0, 12), sticky="ew")

        self.detect_cnpj_var = tk.StringVar(value="CNPJ: -")
        self.detect_mode_var = tk.StringVar(value="Compra para: -")
        self.detect_frete_var = tk.StringVar(value="Frete: -")
        self.total_itens_var = tk.StringVar(value="Itens: 0")
        info_vars = [self.detect_cnpj_var, self.detect_mode_var, self.detect_frete_var, self.total_itens_var]
        for idx, var in enumerate(info_vars, start=2):
            tk.Label(left_card, textvariable=var, anchor="w", bg="#f6f8fc", fg="#16253d", font=("Segoe UI", 10, "bold"), padx=12, pady=10).grid(row=idx, column=0, columnspan=2, sticky="ew", padx=22, pady=(0, 8))

        tk.Label(left_card, text="Formatos aceitos: PDF, XLSX, XLSM, CSV, PNG, JPG e JPEG.", bg="#fff4cc", fg="#5b4a00", font=("Segoe UI", 10), padx=12, pady=10).grid(row=6, column=0, columnspan=2, sticky="ew", padx=22, pady=(8, 18))

        right_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        right_card.grid(row=0, column=1, sticky="nsew")
        right_card.grid_columnconfigure(0, weight=1)
        tk.Label(right_card, text="Como funciona o preenchimento automatico", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=(18, 10))
        proc = (
            "1. Clique em Selecionar arquivo.

"
            "2. Escolha o orcamento em PDF, Excel, CSV ou imagem.

"
            "3. O sistema tenta localizar CNPJ da filial, tipo de frete e tabela de itens.

"
            "4. Cada item identificado e enviado para a grade de resultados abaixo.

"
            "5. Se o arquivo tiver mais de um item, todos sao listados para conferencia.

"
            "6. Ao final, exporte a planilha em Excel formatado."
        )
        tk.Label(right_card, text=proc, justify="left", bg="#eef2ff", fg="#334155", font=("Segoe UI", 11), wraplength=640, padx=18, pady=18).grid(row=1, column=0, sticky="ew", padx=22)

        model = (
            "Cabecalhos recomendados no arquivo: Produto, NCM, IPI, ICMS, Frete, Preco, Quantidade e Codigo.

"
            "Quanto mais estruturado vier o orcamento, melhor a leitura automatica."
        )
        tk.Label(right_card, text=model, justify="left", bg="#fff7e6", fg="#7c5a00", font=("Segoe UI", 10), wraplength=640, padx=16, pady=14).grid(row=2, column=0, sticky="ew", padx=22, pady=(18, 18))

        self._build_results_area(wrapper, row=2)
        self.status_var.set("Modo de anexo ativo. Clique em Selecionar arquivo para anexar o orcamento.")

    def _build_results_area(self, parent, row):
        box = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        box.grid(row=row, column=0, sticky="nsew", pady=(14, 0))
        box.grid_columnconfigure(0, weight=1)
        box.grid_rowconfigure(1, weight=1)

        top = tk.Frame(box, bg="#ffffff")
        top.grid(row=0, column=0, sticky="ew")
        top.grid_columnconfigure(1, weight=1)
        tk.Label(top, text="Resultados", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=14)
        self.item_combo_var = tk.StringVar()
        self.item_combo = ttk.Combobox(top, textvariable=self.item_combo_var, state="readonly")
        self.item_combo.grid(row=0, column=1, sticky="e", padx=22)
        self.item_combo.bind("<<ComboboxSelected>>", self.on_item_selected)

        table_frame = tk.Frame(box, bg="#ffffff")
        table_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        cols = ("idx", "codigo", "produto", "qtde", "ncm", "preco", "icms", "ipi", "frete", "r", "s", "t")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=12)
        headers = {
            "idx": "#", "codigo": "Codigo", "produto": "Produto", "qtde": "Qtde", "ncm": "NCM",
            "preco": "Preco Unitario", "icms": "ICMS", "ipi": "IPI", "frete": "Frete",
            "r": "R / Filial 4", "s": "S / Filial 2", "t": "T / Filial 3 e 5",
        }
        widths = {"idx": 45, "codigo": 90, "produto": 500, "qtde": 60, "ncm": 100, "preco": 115, "icms": 75, "ipi": 75, "frete": 70, "r": 130, "s": 130, "t": 140}
        for col in cols:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor="center")
        self.tree.column("produto", anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.tag_configure("price_highlight", background="#fff4cc")
        self.tree.tag_configure("normal_row", background="#ffffff")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        xscroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

    def limpar_manual(self):
        self.manual_nome.set("")
        self.manual_ncm.set("")
        self.manual_ipi.set("0")
        self.manual_icms.set("7")
        self.manual_preco.set("")
        self.manual_qtde.set("1")
        self.manual_frete.set("FOB")
        self.manual_filial.set(f"Natal - {CNPJ_RN}")
        self.status_var.set("Campos do modo manual limpos.")

    def confirmar_manual(self):
        try:
            nome = self.manual_nome.get().strip()
            ncm = re.sub(r"\D", "", self.manual_ncm.get())
            if not nome:
                raise ValueError("Informe o Nome do Produto.")
            if not ncm:
                raise ValueError("Informe um NCM valido.")

            item = {
                "codigo": "MANUAL",
                "descricao": nome,
                "ncm": ncm,
                "ipi": parse_percent(self.manual_ipi.get()),
                "icms": parse_percent(self.manual_icms.get()),
                "frete": self.manual_frete.get(),
                "preco": to_decimal(self.manual_preco.get()),
                "qtde": int(to_decimal(self.manual_qtde.get()) or Decimal("1")),
            }
            if item["preco"] <= 0:
                raise ValueError("Informe um Preco Unitario maior que zero.")
            if item["qtde"] <= 0:
                item["qtde"] = 1

            self.items = [item]
            compra_para = "RN" if "Natal" in self.manual_filial.get() else "PE"
            if not hasattr(self, "detect_mode_var"):
                self.detect_mode_var = tk.StringVar(value="Compra para: -")
            self.detect_mode_var.set(f"Compra para: {compra_para}")
            self._populate_combo_labels()
            self.calcular_orcamento_inteiro()
            self.status_var.set("Item manual gerado com sucesso.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def _populate_combo_labels(self):
        labels = []
        for i, item in enumerate(self.items, start=1):
            desc = item.get("descricao", "") or "(sem descricao)"
            labels.append(f"{i}. {item.get('codigo','')} | Qtde {item.get('qtde',1)} | {desc[:90]} | NCM: {item.get('ncm','')}")
        if hasattr(self, "item_combo"):
            self.item_combo["values"] = labels
            if labels:
                self.item_combo.current(0)

    def open_budget(self):
        path = filedialog.askopenfilename(
            title="Selecione o arquivo do orcamento",
            filetypes=[
                ("Arquivos suportados", "*.pdf *.xlsx *.xlsm *.csv *.png *.jpg *.jpeg"),
                ("PDF", "*.pdf"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Imagem", "*.png *.jpg *.jpeg"),
            ],
        )
        if not path:
            return
        try:
            data = self.reader.read(path)
            self.budget_path_var.set(path)
            self.detect_cnpj_var.set(f"CNPJ: {data.get('cnpj') or 'nao identificado'}")
            self.detect_mode_var.set(f"Compra para: {data.get('compra_para') or 'nao identificada'}")
            self.detect_frete_var.set(f"Frete: {data.get('frete') or '-'}")
            self.items = data.get("items", [])
            if not self.items:
                raise ValueError("Nao consegui identificar itens automaticamente nesse arquivo.")
            self.total_itens_var.set(f"Itens: {len(self.items)}")
            self._populate_combo_labels()
            self.calcular_orcamento_inteiro()
            self.status_var.set(f"Arquivo lido com sucesso. Itens encontrados: {len(self.items)}.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def on_item_selected(self, event=None):
        idx = self.item_combo.current()
        if idx >= 0 and self.tree.get_children():
            self.tree.selection_set(self.tree.get_children()[idx])
            self.tree.see(self.tree.get_children()[idx])

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        idx = int(values[0]) - 1
        if 0 <= idx < len(self.items):
            self.item_combo.current(idx)

    def calcular_orcamento_inteiro(self):
        if not self.items:
            messagebox.showwarning(APP_TITLE, "Nao ha itens para calcular.")
            return

        if hasattr(self, "tree"):
            for item_id in self.tree.get_children():
                self.tree.delete(item_id)

        self.current_results = []
        compra_para = ""
        cnpj_text = self.detect_mode_var.get().upper() if hasattr(self, "detect_mode_var") else ""
        if "RN" in cnpj_text:
            compra_para = "RN"
        elif "PE" in cnpj_text:
            compra_para = "PE"
        elif len(self.items) == 1 and hasattr(self, "manual_filial"):
            compra_para = "RN" if "Natal" in self.manual_filial.get() else "PE"

        if compra_para not in ("RN", "PE"):
            raise ValueError("Nao foi possivel identificar automaticamente RN ou PE.")

        for idx, item in enumerate(self.items, start=1):
            try:
                result = self.engine.calcular(compra_para, item["preco"], item["ncm"], item["icms"], item["ipi"], item["frete"])
                record = {
                    "idx": idx,
                    "codigo": item.get("codigo", ""),
                    "produto": item["descricao"],
                    "qtde": item["qtde"],
                    "ncm": item["ncm"],
                    "preco": item["preco"],
                    "icms": item["icms"],
                    "ipi": item["ipi"],
                    "frete": item["frete"],
                    "r": result["r"],
                    "s": result["s"],
                    "t": result["t"],
                }
                self.current_results.append(record)
                self.tree.insert(
                    "", "end",
                    values=(record["idx"], record["codigo"], record["produto"], record["qtde"], record["ncm"], format_money(record["preco"]), format_pct(record["icms"]), format_pct(record["ipi"]), record["frete"], format_money(record["r"]), format_money(record["s"]), format_money(record["t"])),
                    tags=("price_highlight",)
                )
            except Exception as e:
                self.tree.insert(
                    "", "end",
                    values=(idx, item.get("codigo", ""), f"ERRO: {str(e)}", item.get("qtde", 1), item.get("ncm", ""), "", "", "", item.get("frete", ""), "", "", ""),
                    tags=("normal_row",)
                )
        self.status_var.set(f"Precificacao concluida. Total de itens: {len(self.items)}.")

    def exportar_excel(self):
        if not self.current_results:
            messagebox.showwarning(APP_TITLE, "Calcule pelo menos um item antes de exportar.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Salvar resultado em Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="resultado_precificacao_profissional.xlsx",
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultado"
        headers = ["Item", "Codigo", "Produto", "Qtde", "NCM", "Preco Unitario", "ICMS", "IPI", "Frete", "Resultado R", "Resultado S", "Resultado T"]
        ws.append(headers)
        for rec in self.current_results:
            ws.append([
                rec["idx"], rec["codigo"], rec["produto"], rec["qtde"], rec["ncm"],
                float(q2(rec["preco"])), float(rec["icms"]), float(rec["ipi"]), rec["frete"],
                float(q2(rec["r"])), float(q2(rec["s"])), float(q2(rec["t"])),
            ])

        header_fill = PatternFill(fill_type="solid", fgColor="6F2DBD")
        header_font = Font(color="FFFFFF", bold=True)
        price_fill = PatternFill(fill_type="solid", fgColor="FFF4CC")
        border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        money_cols = {"F", "J", "K", "L"}
        pct_cols = {"G", "H"}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column_letter in money_cols and cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'
                elif cell.column_letter in pct_cols and cell.value is not None:
                    cell.number_format = '0.00%'
                if cell.column_letter == "F":
                    cell.fill = price_fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 70)

        ws.freeze_panes = "A2"
        wb.save(filepath)
        messagebox.showinfo(APP_TITLE, f"Arquivo exportado com sucesso:\n{filepath}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
