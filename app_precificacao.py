import os
import re
import csv
import json
import unicodedata
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_TITLE = "Assistente Profissional de Orcamento"
DEFAULT_WORKBOOK = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "1 NOVO SIMULADOR_PRECIFICACAO_V2.xlsx",
)

PIS_COFINS_RATE = Decimal("0.0925")
FRETE_FOB_RATE_RN = Decimal("0.10")
FRETE_FOB_RATE_PE = Decimal("0.05")
ROUND_MONEY = Decimal("0.01")

CNPJ_RN = "18.217.682/0004-05"
CNPJ_PE = "18.217.682/0001-54"
CNPJ_RN_DIGITS = "18217682000405"
CNPJ_PE_DIGITS = "18217682000154"


def q2(value: Decimal) -> Decimal:
    return value.quantize(ROUND_MONEY, rounding=ROUND_HALF_UP)


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip().replace("R$", "").replace("%", "").replace(" ", "")
    if text in ("-", "—", "None"):
        return Decimal("0")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return Decimal(text)


def parse_percent(value) -> Decimal:
    v = to_decimal(value)
    if v > Decimal("1"):
        v = v / Decimal("100")
    return v


def format_money(value: Decimal) -> str:
    value = q2(value)
    s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def format_pct(value: Decimal) -> str:
    v = (value * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}%"


def strip_accents(text: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(text)) if not unicodedata.combining(ch))


def normalize_text(text: str) -> str:
    base = strip_accents(str(text)).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", base)


class PricingEngine:
    def __init__(self, workbook_path: str):
        self.ncm_factors = {}
        self.st_rates = {}
        self.load_workbook(workbook_path)

    def load_workbook(self, workbook_path: str):
        if not os.path.exists(workbook_path):
            raise FileNotFoundError(
                f"Planilha base nao encontrada.\n\n"
                f"Coloque o arquivo '{os.path.basename(workbook_path)}' na mesma pasta do programa."
            )

        wb = openpyxl.load_workbook(workbook_path, data_only=False)

        ws_ncm = wb["Base Dados NCM"]
        self.ncm_factors.clear()
        for row in ws_ncm.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.ncm_factors[key] = {
                "fator_custo_filial_35": to_decimal(row[1]),
                "fator_psd_filial_35": to_decimal(row[3]),
                "fator_psd_filial_2_fora_pe": to_decimal(row[4]),
                "fator_psd_filial_4_dentro_rn": to_decimal(row[5]),
                "fator_psd_filial_4_fora_rn": to_decimal(row[6]),
            }

        ws_st = wb["Base ST RN"]
        self.st_rates.clear()
        for row in ws_st.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.st_rates[key] = to_decimal(row[1])

    def calcular(
        self,
        compra_para: str,
        preco_unitario: Decimal,
        ncm: str,
        icms: Decimal,
        ipi: Decimal,
        frete_tipo: str,
    ):
        ncm_key = str(ncm).strip()
        if ncm_key not in self.ncm_factors:
            raise KeyError(f"NCM {ncm_key} nao encontrado na aba 'Base Dados NCM'.")

        fatores = self.ncm_factors[ncm_key]
        st_rate = self.st_rates.get(ncm_key, Decimal("0"))

        vlr_icms = preco_unitario * icms
        vlr_ipi = preco_unitario * ipi
        bc_pis_cofins = preco_unitario * (Decimal("1") - icms)
        pis_cofins = (bc_pis_cofins + vlr_ipi) * PIS_COFINS_RATE

        if compra_para == "RN":
            frete_fob = preco_unitario * FRETE_FOB_RATE_RN if frete_tipo == "FOB" else Decimal("0")
            vlr_st = preco_unitario * st_rate
            custo_filial_4 = (preco_unitario + vlr_ipi + vlr_st + frete_fob) - pis_cofins
            custo_filial_2 = custo_filial_4 * Decimal("0.88")
            custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
            r = custo_filial_4 * fatores["fator_psd_filial_4_dentro_rn"]
            s = custo_filial_4 * fatores["fator_psd_filial_4_fora_rn"]
            t = custo_filial_35 * fatores["fator_psd_filial_35"]
            return {"r": r, "s": s, "t": t}

        frete_fob = (preco_unitario + vlr_ipi) * FRETE_FOB_RATE_PE if frete_tipo == "FOB" else Decimal("0")
        custo_filial_2 = (preco_unitario + vlr_ipi + frete_fob) - pis_cofins - vlr_icms
        custo_filial_4 = custo_filial_2 * Decimal("1.12")
        custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
        r = custo_filial_4
        s = custo_filial_2 * fatores["fator_psd_filial_2_fora_pe"]
        t = custo_filial_35 * fatores["fator_psd_filial_35"]
        return {"r": r, "s": s, "t": t}


class UniversalBudgetReader:
    HEADER_ALIASES = {
        "descricao": ["descricao", "descrição", "produto", "item", "nomeproduto", "descricaoproduto", "material"],
        "ncm": ["ncm", "ncmsh", "codigoncm", "classfiscal", "classificacaofiscal", "classificaçãofiscal"],
        "ipi": ["ipi", "aliquotaipi", "percipi", "percentualipi"],
        "icms": ["icms", "aliquotaicms", "percicms", "percentualicms"],
        "frete": ["frete", "tipofrete", "modofrete", "fobcif", "ciffob", "tipo frete"],
        "preco": ["preco", "preço", "precounitario", "preçounitário", "valorunitario", "vlunitario", "valor", "unitario", "unitário", "preco rsun", "valorunitario", "valor unitario"],
        "codigo": ["codigo", "código", "cod", "referencia", "referência"],
        "qtde": ["qtde", "quantidade", "qtd", "qt"],
    }

    def __init__(self):
        self.diagnostics = []
        self.last_strategy = ""
        self.learning_enabled = True
        self.learning_db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "orcamento_learning_db.json")
        self.learning_db = self._load_learning_db()

    def _load_learning_db(self):
        if not os.path.exists(self.learning_db_path):
            return {"profiles": {}}
        try:
            with open(self.learning_db_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return {"profiles": {}}
            data.setdefault("profiles", {})
            return data
        except Exception:
            return {"profiles": {}}

    def _save_learning_db(self):
        try:
            with open(self.learning_db_path, "w", encoding="utf-8") as f:
                json.dump(self.learning_db, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._log(f"Falha ao salvar base de aprendizado: {e}")

    def reset_diagnostics(self):
        self.diagnostics = []
        self.last_strategy = ""

    def _log(self, message: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.diagnostics.append(f"[{ts}] {message}")

    def get_diagnostics_text(self):
        if not self.diagnostics:
            return "Nenhum diagnóstico disponível."
        return "\n".join(self.diagnostics)

    def _extract_cnpjs(self, text_dump: str):
        nums = re.sub(r"\D", "", text_dump)
        return list(dict.fromkeys(re.findall(r"\d{14}", nums)))

    def _guess_supplier_name(self, text_dump: str, filepath: str = ""):
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in text_dump.splitlines() if ln.strip()]
        ranked = []
        for line in lines[:80] + lines[-40:]:
            norm = strip_accents(line).upper()
            if any(term in norm for term in [" LTDA", " S/A", " EIRELI", " COMERCIO", " DISTRIBUIDORA", " INDUSTRIA", " IND.", " CABOS", " SOLUCOES", " TECNOLOGIA"]):
                if "KGMLAN" in norm or CNPJ_RN_DIGITS in re.sub(r"\D", "", norm) or CNPJ_PE_DIGITS in re.sub(r"\D", "", norm):
                    continue
                ranked.append(line[:90])
        if ranked:
            return ranked[0]
        if filepath:
            return os.path.splitext(os.path.basename(filepath))[0]
        return "layout_desconhecido"

    def _profile_key(self, text_dump: str, filepath: str = ""):
        supplier = normalize_text(self._guess_supplier_name(text_dump, filepath))[:80] or "layout_desconhecido"
        cnpjs = self._extract_cnpjs(text_dump)
        issuer = ""
        for cnpj in cnpjs:
            if cnpj not in (CNPJ_RN_DIGITS, CNPJ_PE_DIGITS):
                issuer = cnpj
                break
        return f"{supplier}__{issuer}" if issuer else supplier

    def _learn_layout(self, filepath: str, text_dump: str, items, strategy: str):
        if not self.learning_enabled:
            self._log("Aprendizado desativado para esta leitura.")
            return
        key = self._profile_key(text_dump, filepath)
        supplier = self._guess_supplier_name(text_dump, filepath)
        profile = self.learning_db.setdefault("profiles", {}).setdefault(key, {
            "supplier_name": supplier,
            "first_seen": datetime.now().isoformat(timespec="seconds"),
            "successful_reads": 0,
            "strategies": {},
            "sample_headers": [],
            "known_cnpjs": [],
        })
        profile["last_seen"] = datetime.now().isoformat(timespec="seconds")
        profile["successful_reads"] = int(profile.get("successful_reads", 0)) + 1
        profile.setdefault("strategies", {})
        profile["strategies"][strategy] = int(profile["strategies"].get(strategy, 0)) + 1
        profile["last_strategy"] = strategy
        profile["last_item_count"] = len(items)
        profile["known_cnpjs"] = list(dict.fromkeys(profile.get("known_cnpjs", []) + self._extract_cnpjs(text_dump)))[:10]

        sample_headers = []
        for line in [ln.strip() for ln in text_dump.splitlines() if ln.strip()][:40]:
            compact = re.sub(r"\s+", " ", line)
            if sum(1 for field in self.HEADER_ALIASES if normalize_text(field) in normalize_text(compact)) >= 1:
                sample_headers.append(compact[:120])
        if sample_headers:
            existing = profile.get("sample_headers", [])
            profile["sample_headers"] = list(dict.fromkeys(existing + sample_headers))[:20]

        self.learning_db["profiles"][key] = profile
        self._save_learning_db()
        self._log(f"Aprendizado salvo para o layout '{supplier}' usando estratégia '{strategy}'.")

    def get_learning_summary(self):
        profiles = self.learning_db.get("profiles", {})
        if not profiles:
            return "Nenhum layout aprendido até o momento."
        linhas = ["Modelos aprendidos:"]
        for _, profile in sorted(profiles.items(), key=lambda kv: kv[1].get("last_seen", ""), reverse=True)[:20]:
            linhas.append(f"- {profile.get('supplier_name', 'layout')} | leituras: {profile.get('successful_reads', 0)} | última estratégia: {profile.get('last_strategy', '-')}")
        return "\n".join(linhas)

    def read(self, filepath: str, learning_enabled: bool = True):
        self.reset_diagnostics()
        self.learning_enabled = learning_enabled
        ext = os.path.splitext(filepath)[1].lower()
        self._log(f"Iniciando leitura do arquivo: {os.path.basename(filepath)}")
        self._log(f"Extensão detectada: {ext}")
        if ext in (".xlsx", ".xlsm"):
            data = self._read_excel(filepath)
        elif ext == ".csv":
            data = self._read_csv(filepath)
        elif ext == ".pdf":
            data = self._read_pdf(filepath)
        elif ext in (".png", ".jpg", ".jpeg"):
            data = self._read_image(filepath)
        else:
            raise ValueError("Formato nao suportado. Use PDF, XLSX, XLSM, CSV, PNG, JPG ou JPEG.")

        data["diagnostics_text"] = self.get_diagnostics_text()
        data["learning_summary"] = self.get_learning_summary()
        return data

    def _detect_cnpj_and_mode(self, text_dump: str):
        texto_numeros = re.sub(r"\D", "", text_dump)
        cnpjs = re.findall(r"\d{14}", texto_numeros)
        self._log(f"CNPJs encontrados no texto: {len(cnpjs)}")
        for cnpj in cnpjs:
            cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            if cnpj == CNPJ_RN_DIGITS or cnpj.endswith("000405"):
                self._log(f"CNPJ de destino reconhecido como RN: {cnpj_formatado}")
                return cnpj_formatado, "RN"
            if cnpj == CNPJ_PE_DIGITS or cnpj.endswith("000154"):
                self._log(f"CNPJ de destino reconhecido como PE: {cnpj_formatado}")
                return cnpj_formatado, "PE"
        upper = text_dump.upper()
        if "KGMLAN" in upper or "KGM LAN" in upper:
            self._log("Nome KGMLAN detectado sem CNPJ explícito. Assumindo RN.")
            return CNPJ_RN, "RN"
        self._log("Não foi possível identificar automaticamente RN/PE pelo CNPJ.")
        return "", ""

    def _detect_frete(self, text_dump: str):
        upper = strip_accents(text_dump).upper()
        explicit_tipo = re.search(r"TIPO\s+FRETE\s+([A-Z\.]+)", upper)
        if explicit_tipo:
            valor = re.sub(r"[^A-Z]", "", explicit_tipo.group(1))
            if valor == "FOB":
                self._log("Frete identificado pela expressão 'TIPO FRETE': FOB")
                return "FOB"
            if valor == "CIF":
                self._log("Frete identificado pela expressão 'TIPO FRETE': CIF")
                return "CIF"
        if re.search(r"\bC\.?I\.?F\.?\b", upper):
            self._log("Frete identificado como CIF.")
            return "CIF"
        if re.search(r"\bF\.?O\.?B\.?\b", upper):
            self._log("Frete identificado como FOB.")
            return "FOB"
        compact = re.sub(r"[^A-Z0-9]+", " ", upper)
        if "FRETE CIF" in compact:
            self._log("Frete identificado pelo texto compacto: CIF")
            return "CIF"
        if "FRETE FOB" in compact:
            self._log("Frete identificado pelo texto compacto: FOB")
            return "FOB"
        self._log("Frete não identificado com clareza. Assumindo CIF como padrão.")
        return "CIF"

    def _safe_decimal(self, value, default="0"):
        try:
            return to_decimal(value)
        except Exception:
            return Decimal(default)

    def _extract_default_tax_rates(self, text_dump: str):
        txt = strip_accents(text_dump).upper()
        total_produtos = None
        total_ipi = None
        total_icms = None

        product_patterns = [
            r"VALOR\s+TOTAL\s+DOS\s+PRODUTOS\s*[:\-]?\s*([\d\.,]+)",
            r"MERCADORIA\s*[:\-]?\s*([\d\.,]+)",
            r"TOTAL\s+PRODUTOS\s*[:\-]?\s*([\d\.,]+)",
        ]
        for pat in product_patterns:
            m_prod = re.search(pat, txt)
            if m_prod:
                total_produtos = self._safe_decimal(m_prod.group(1))
                self._log(f"Base de produtos inferida pelo padrão '{pat}': {m_prod.group(1)}")
                break

        valores_block = None
        vals = re.findall(r"VALORES?\s+DO\s+ORCAMENTO.*?([\d\.,\s]{20,})", txt, flags=re.S)
        if vals:
            valores_block = vals[0]
            nums = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", valores_block)
            if (total_produtos is None or total_produtos == 0) and nums:
                total_produtos = self._safe_decimal(nums[-2] if len(nums) >= 2 else nums[-1])
                self._log(f"Base de produtos inferida pelo bloco 'Valores do Orçamento': {total_produtos}")

        ipi_patterns = [
            r"VALOR\s+DO\s+IPI\s*[:\-]?\s*([\d\.,]+)",
            r"\bIPI\s*\.?\s*[:\-]?\s*([\d\.,]+)",
        ]
        for pat in ipi_patterns:
            m_ipi = re.search(pat, txt)
            if m_ipi:
                total_ipi = self._safe_decimal(m_ipi.group(1))
                self._log(f"Total de IPI inferido pelo padrão '{pat}': {m_ipi.group(1)}")
                break

        icms_patterns = [
            r"VALOR\s+ICMS\s*[:\-]?\s*([\d\.,]+)",
            r"\bICMS\s*\.?\s*[:\-]?\s*([\d\.,]+)",
        ]
        icms_matches = []
        for pat in icms_patterns:
            icms_matches.extend(re.findall(pat, txt))
        if icms_matches:
            positivos = [self._safe_decimal(v) for v in icms_matches if self._safe_decimal(v) > 0]
            if positivos:
                total_icms = max(positivos)
                self._log(f"Total de ICMS inferido pelos padrões textuais: {total_icms}")

        if (total_icms is None or total_icms == 0) and valores_block:
            nums = [self._safe_decimal(v) for v in re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", valores_block)]
            positivos = [n for n in nums if n > 0]
            if len(positivos) >= 2:
                total_icms = positivos[1]
                self._log(f"Total de ICMS inferido pelo bloco 'Valores do Orçamento': {total_icms}")

        icms = Decimal("0")
        ipi = Decimal("0")
        if total_produtos and total_produtos > 0:
            if total_icms and total_icms > 0:
                icms = (total_icms / total_produtos).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
            if total_ipi and total_ipi > 0:
                ipi = (total_ipi / total_produtos).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

        self._log(f"Tributos padrão inferidos do orçamento -> ICMS: {format_pct(icms)}, IPI: {format_pct(ipi)}")
        return {"icms": icms, "ipi": ipi}


    def _match_header(self, value):
        normalized = normalize_text(value)
        for field, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                if normalize_text(alias) == normalized:
                    return field
        return None

    def _merge_learned_aliases(self, profile_key: str):
        profile = self.learning_db.get("profiles", {}).get(profile_key)
        if not profile:
            return
        self._log(f"Perfil aprendido carregado para '{profile.get('supplier_name', profile_key)}'.")

    def _dedupe_items(self, items):
        deduped = []
        seen = set()
        for item in items:
            key = (
                str(item.get("codigo", "")).strip(),
                str(item.get("descricao", "")).strip().upper(),
                re.sub(r"\D", "", str(item.get("ncm", ""))),
                str(q2(to_decimal(item.get("preco", "")))),
                str(item.get("qtde", "")),
            )
            if key in seen:
                self._log(f"Item duplicado descartado: {item.get('descricao', '')[:60]}")
                continue
            seen.add(key)
            deduped.append(item)
        return deduped

    def _post_process_items(self, items, default_frete):
        processed = []
        for item in self._dedupe_items(items):
            desc = str(item.get("descricao", "")).strip()
            ncm = re.sub(r"\D", "", str(item.get("ncm", "")))
            preco = to_decimal(item.get("preco", ""))
            qtde_dec = to_decimal(item.get("qtde", "1"))
            qtde = int(qtde_dec) if qtde_dec > 0 else 1
            motivos = []
            if qtde <= 0:
                motivos.append("quantidade <= 0")
            if preco <= 0:
                motivos.append("preço <= 0")
            if not desc and not ncm:
                motivos.append("sem descrição e sem NCM")
            if motivos:
                self._log(f"Item descartado no pós-processamento: {', '.join(motivos)}")
                continue
            processed.append({
                "codigo": str(item.get("codigo", "")).strip(),
                "descricao": desc,
                "ncm": ncm,
                "ipi": parse_percent(item.get("ipi", "")),
                "icms": parse_percent(item.get("icms", "")),
                "frete": self._detect_frete(str(item.get("frete", ""))) or default_frete,
                "preco": preco,
                "qtde": qtde,
            })
        self._log(f"Itens válidos após pós-processamento: {len(processed)}")
        return processed

    def _extract_from_tabular_rows(self, rows, text_dump, context="tabela"):
        found_items = []
        detected_headers = {}
        for idx, row in enumerate(rows[:30]):
            current_map = {}
            for col_idx, val in enumerate(row):
                if val in (None, ""):
                    continue
                matched = self._match_header(val)
                if matched:
                    current_map[matched] = col_idx
            if current_map:
                detected_headers = current_map
            if ("descricao" in current_map and "ncm" in current_map and "preco" in current_map and "qtde" in current_map):
                self._log(f"Cabeçalhos reconhecidos em {context}: {sorted(current_map.keys())}")
                for data_row in rows[idx + 1:]:
                    if not any(v not in (None, "") for v in data_row):
                        continue
                    item = {}
                    for field, col_idx in current_map.items():
                        if col_idx < len(data_row):
                            item[field] = data_row[col_idx]
                    found_items.append(item)
                break
        if found_items:
            self.last_strategy = "headers_tabulares"
            self._log(f"Extração tabular encontrou {len(found_items)} item(ns).")
            return found_items

        if detected_headers:
            self._log(f"Cabeçalhos parciais encontrados em {context}, mas insuficientes: {sorted(detected_headers.keys())}")
        else:
            self._log(f"Nenhum cabeçalho estruturado útil encontrado em {context}.")

        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        generic_items = []
        i = 0
        while i < len(lines):
            line = lines[i]
            m_head = re.match(r'^([A-Z0-9\-]+)\s+([\d\.,]+)\s+(\d+)\s+([\d\.,]+)(?:\s+[\d\.]+)?$', line)
            if m_head:
                codigo, preco, qtde, _total = m_head.groups()
                desc_lines = []
                j = i + 1
                while j < len(lines) and not re.search(r'\b\d{8}\b', lines[j]):
                    desc_lines.append(lines[j])
                    j += 1
                if j < len(lines):
                    tail = lines[j]
                    m_tail = re.search(r'(\d{8}).*?(\d{1,2}(?:[.,]\d{1,2})?).*?(\d{1,2}(?:[.,]\d{1,2})?)', tail)
                    if m_tail:
                        ncm, icms, ipi = m_tail.groups()
                        generic_items.append({
                            "codigo": codigo,
                            "descricao": " ".join(desc_lines).replace("#", "").strip(),
                            "ncm": ncm,
                            "icms": icms,
                            "ipi": ipi,
                            "preco": preco,
                            "frete": self._detect_frete(text_dump),
                            "qtde": qtde,
                        })
                        i = j
            i += 1
        if generic_items:
            self.last_strategy = "fallback_linhas_genericas"
            self._log(f"Fallback por linhas genéricas encontrou {len(generic_items)} item(ns).")
        else:
            self._log("Fallback por linhas genéricas não encontrou itens.")
        return generic_items

    def _extract_items_from_text_dump(self, text_dump):
        default_taxes = self._extract_default_tax_rates(text_dump)
        default_frete = self._detect_frete(text_dump)
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in text_dump.splitlines() if ln.strip()]
        found = []

        row_patterns = [
            ("linha_compacta_completa", re.compile(r"^(?P<idx>\d{1,3})\s+(?P<codigo>[A-Z0-9\-]{2,})\s+(?P<descricao>.+?)\s+(?P<qtde>\d+)\s+(?P<un>[A-Z]{2,4})\s+(?P<preco>\d{1,3}(?:\.\d{3})*,\d{2,4})\s+(?P<total>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<ipi>\d{1,2}(?:,\d{1,2})?)\s+(?P<mva>\d{1,2}(?:,\d{1,2})?)\s+(?P<ncm>\d{8})$")),
            ("linha_compacta_icms_ipi", re.compile(r"^(?P<codigo>[A-Z0-9\-]{2,})\s+(?P<descricao>.+?)\s+(?P<qtde>\d+)\s+(?P<preco>\d{1,3}(?:\.\d{3})*,\d{2,4})\s+(?P<total>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<icms>\d{1,2}(?:,\d{1,2})?)\s+(?P<ipi>\d{1,2}(?:,\d{1,2})?)\s+(?P<ncm>\d{8})$")),
            ("linha_megatron_unica", re.compile(r"^(?P<codigo>[A-Z0-9\-]{6,})\s+(?:\((?P<codigo_alt>[^\)]+)\)\s+)?(?P<qtde>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s+(?P<un>[A-Z]{2,4})\s+(?P<descricao>.+?)\s+(?P<metragem>\d{1,3}(?:\.\d{3})*,\d{2})\s+[A-Z]{1,3}\s+(?P<preco>\d{1,3}(?:\.\d{3})*,\d{4})\s+(?P<mercadoria>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<st>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<ipi_valor>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<total>\d{1,3}(?:\.\d{3})*,\d{2})$")),
        ]

        for line in lines:
            for strategy_name, pattern in row_patterns:
                m = pattern.match(line)
                if not m:
                    continue
                gd = m.groupdict()
                found.append({
                    "codigo": gd.get("codigo", ""),
                    "descricao": gd.get("descricao", "").strip(" -"),
                    "qtde": gd.get("qtde", "1"),
                    "preco": gd.get("preco", "0"),
                    "ncm": gd.get("ncm", ""),
                    "icms": gd.get("icms") or default_taxes["icms"],
                    "ipi": gd.get("ipi") or default_taxes["ipi"],
                    "frete": default_frete,
                })
                self.last_strategy = strategy_name
                break

        if found:
            self._log(f"Extração por padrões de linha encontrou {len(found)} item(ns) com estratégia '{self.last_strategy}'.")
            return found

        detail_pattern = re.compile(r"^(?P<qtde>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s+(?P<un>[A-Z]{2,4})\s+(?P<descricao>.+?)\s+(?P<metragem>\d{1,3}(?:\.\d{3})*,\d{2})\s+[A-Z]{1,3}\s+(?P<preco>\d{1,3}(?:\.\d{3})*,\d{4})\s+(?P<mercadoria>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<st>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<ipi_valor>\d{1,3}(?:\.\d{3})*,\d{2})\s+(?P<total>\d{1,3}(?:\.\d{3})*,\d{2})$")
        for i in range(len(lines) - 1):
            codigo = ""
            detail_line = ""
            search_start = i + 1

            if re.match(r"^[A-Z0-9\-]{6,}$", lines[i]):
                codigo = lines[i]
                if i + 2 < len(lines) and re.match(r"^\([^\)]+\)$", lines[i + 1]):
                    detail_line = lines[i + 2]
                    search_start = i + 3
                else:
                    detail_line = lines[i + 1]
                    search_start = i + 2
            elif re.match(r"^[A-Z0-9\-]{6,}\s+\([^\)]+\)$", lines[i]):
                codigo = lines[i].split()[0]
                detail_line = lines[i + 1]
                search_start = i + 2
            else:
                continue

            m = detail_pattern.match(detail_line)
            if not m:
                continue

            gd = m.groupdict()
            trailing_ncm = ""
            for extra in lines[search_start:search_start + 4]:
                m_ncm = re.search(r"\b(\d{8})\b", extra)
                if m_ncm:
                    trailing_ncm = m_ncm.group(1)
                    break

            found.append({
                "codigo": codigo,
                "descricao": gd.get("descricao", "").strip(" -"),
                "qtde": gd.get("qtde", "1"),
                "preco": gd.get("preco", "0"),
                "ncm": trailing_ncm,
                "icms": default_taxes["icms"],
                "ipi": default_taxes["ipi"],
                "frete": default_frete,
            })
            self.last_strategy = "linha_megatron_multilinha"

        if found:
            self._log(f"Extração multilinha encontrou {len(found)} item(ns) com estratégia '{self.last_strategy}'.")
            return found

        i = 0
        while i < len(lines):
            line = lines[i]
            nums = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2,4}|\b\d{8}\b", line)
            if len(nums) >= 4 and any(len(n) == 8 and n.isdigit() for n in nums):
                ncm_match = re.search(r"(\d{8})\s*$", line)
                qtd_match = re.search(r"\s(\d+)\s+[A-Z]{2,4}\s+\d{1,3}(?:\.\d{3})*,\d{2,4}", line)
                codigo_match = re.match(r"^(?:\d{1,3}\s+)?([A-Z0-9\-]{2,})\s+", line)
                if ncm_match and qtd_match and codigo_match:
                    ncm = ncm_match.group(1)
                    qtd = qtd_match.group(1)
                    codigo = codigo_match.group(1)
                    preco_candidates = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2,4}", line)
                    preco = preco_candidates[0] if preco_candidates else "0"
                    desc_part = re.sub(r"^(?:\d{1,3}\s+)?" + re.escape(codigo) + r"\s+", "", line)
                    desc_part = re.sub(r"\s+" + re.escape(qtd) + r"\s+[A-Z]{2,4}.*$", "", desc_part).strip()
                    if self._looks_like_product_description(desc_part):
                        found.append({
                            "codigo": codigo,
                            "descricao": desc_part,
                            "qtde": qtd,
                            "preco": preco,
                            "ncm": ncm,
                            "icms": default_taxes["icms"],
                            "ipi": default_taxes["ipi"],
                            "frete": default_frete,
                        })
            i += 1

        if found:
            self.last_strategy = "heuristica_texto_livre"
            self._log(f"Heurística de texto livre encontrou {len(found)} item(ns).")
        else:
            self._log("Heurística de texto livre não encontrou itens.")
        return found


    def _finalize_result(self, filepath: str, text_dump: str, found_items, cnpj, compra_para, frete, source_type: str):
        processed = self._post_process_items(found_items, frete)
        strategy = self.last_strategy or "nao_definida"
        if processed:
            self._log(f"Leitura concluída com sucesso usando estratégia '{strategy}'.")
            self._learn_layout(filepath, text_dump, processed, strategy)
        else:
            self._log("Leitura concluída sem itens válidos. Verifique o diagnóstico para entender o motivo.")
        return {
            "cnpj": cnpj,
            "compra_para": compra_para,
            "frete": frete,
            "items": processed,
            "source_type": source_type,
            "strategy": strategy,
            "supplier_name": self._guess_supplier_name(text_dump, filepath),
        }

    def _read_excel(self, filepath: str):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        text_parts, found_items = [], []
        self._log(f"Lendo workbook com {len(wb.worksheets)} aba(s).")
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            sheet_text = []
            for row in rows:
                row_text = " ".join([str(v) for v in row if v not in (None, "")])
                if row_text:
                    sheet_text.append(row_text)
                    text_parts.append(row_text)
            found_items.extend(self._extract_from_tabular_rows(rows, "\n".join(sheet_text), context=f"aba {ws.title}"))
        text_dump = "\n".join(text_parts)
        profile_key = self._profile_key(text_dump, filepath)
        self._merge_learned_aliases(profile_key)
        if not found_items:
            found_items.extend(self._extract_items_from_text_dump(text_dump))
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return self._finalize_result(filepath, text_dump, found_items, cnpj, compra_para, frete, "excel")

    def _read_csv(self, filepath: str):
        with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        text_dump = "\n".join([" ".join([str(v) for v in row if v not in (None, "")]) for row in rows])
        profile_key = self._profile_key(text_dump, filepath)
        self._merge_learned_aliases(profile_key)
        found_items = self._extract_from_tabular_rows(rows, text_dump, context="csv")
        if not found_items:
            found_items.extend(self._extract_items_from_text_dump(text_dump))
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return self._finalize_result(filepath, text_dump, found_items, cnpj, compra_para, frete, "csv")

    def _read_pdf(self, filepath: str):
        try:
            import pdfplumber
        except Exception:
            raise ImportError("Para ler PDF, instale pdfplumber: python -m pip install pdfplumber")
        text_parts, found_items = [], []
        with pdfplumber.open(filepath) as pdf:
            self._log(f"PDF com {len(pdf.pages)} página(s).")
            for page_number, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                text_parts.append(text)
                self._log(f"Página {page_number}: {len(text.splitlines())} linha(s) de texto extraídas.")
                tables = page.extract_tables() or []
                self._log(f"Página {page_number}: {len(tables)} tabela(s) detectada(s).")
                for table_idx, table in enumerate(tables, start=1):
                    if table:
                        found_items.extend(self._extract_from_tabular_rows(table, text, context=f"pdf página {page_number} tabela {table_idx}"))
        text_dump = "\n".join(text_parts)
        profile_key = self._profile_key(text_dump, filepath)
        self._merge_learned_aliases(profile_key)
        if not found_items:
            self._log("Partindo para fallback textual do PDF.")
            found_items.extend(self._extract_items_from_text_dump(text_dump))
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return self._finalize_result(filepath, text_dump, found_items, cnpj, compra_para, frete, "pdf")

    def _read_image(self, filepath: str):
        try:
            from PIL import Image
            import pytesseract
        except Exception:
            raise ImportError("Para ler imagem, instale Pillow e pytesseract, alem do Tesseract OCR no Windows.")
        image = Image.open(filepath)
        text_dump = pytesseract.image_to_string(image, lang="por+eng", config="--psm 6")
        self._log(f"OCR executado na imagem com {len(text_dump.splitlines())} linha(s) reconhecidas.")
        profile_key = self._profile_key(text_dump, filepath)
        self._merge_learned_aliases(profile_key)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        found_items = self._extract_items_from_text_dump(text_dump)
        if not found_items:
            lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
            current = {"codigo": "", "descricao": "", "ncm": "", "icms": "", "ipi": "", "preco": "", "frete": frete, "qtde": "1"}
            for line in lines:
                upper = line.upper()
                if not current["descricao"] and not any(x in upper for x in ["NCM", "ICMS", "IPI", "CNPJ", "R$"]):
                    current["descricao"] = line
                if not current["ncm"]:
                    m = re.search(r"\b(\d{8})\b", line)
                    if m:
                        current["ncm"] = m.group(1)
                if not current["icms"] and "ICMS" in upper:
                    p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                    if p:
                        current["icms"] = p.group(1)
                if not current["ipi"] and "IPI" in upper:
                    p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                    if p:
                        current["ipi"] = p.group(1)
                if not current["preco"] and ("R$" in upper or "UNIT" in upper or "VALOR" in upper or "PRECO" in upper or "PREÇO" in upper):
                    m = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})', line)
                    if m:
                        current["preco"] = m.group(1)
                if current["qtde"] == "1":
                    m = re.search(r'\bQTD(?:E)?\s*:?\s*(\d+)\b', upper)
                    if m:
                        current["qtde"] = m.group(1)
            if current["descricao"] or current["ncm"]:
                found_items.append(current)
                self.last_strategy = "ocr_item_unico"
                self._log("OCR fallback montou 1 item a partir de campos soltos.")
        return self._finalize_result(filepath, text_dump, found_items, cnpj, compra_para, frete, "imagem")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1520x930")
        self.minsize(1380, 860)
        self.configure(bg="#eceff5")

        self.engine = PricingEngine(DEFAULT_WORKBOOK)
        self.reader = UniversalBudgetReader()
        self.items = []
        self.current_results = []
        self.budget_data = None

        self.psd_r_var = tk.StringVar(value="R / Filial 4: -")
        self.psd_s_var = tk.StringVar(value="S / Filial 4 Fora RN / Filial 2 Fora PE: -")
        self.psd_t_var = tk.StringVar(value="T / Filial 3 e 5: -")

        self.attach_filial_manual = tk.StringVar(value=f"Natal - {CNPJ_RN}")
        self.attach_learning_enabled = tk.BooleanVar(value=True)

        self._build_styles()
        self._build_layout()
        self.show_home()

    def _build_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except Exception:
            pass
        style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=(14, 10))
        style.configure("Ghost.TButton", font=("Segoe UI", 10), padding=(12, 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30)

    def _build_layout(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        header = tk.Frame(self, bg="#6f2dbd", height=110)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)
        tk.Label(
            header,
            text="ASSISTENTE PROFISSIONAL DE ORCAMENTO",
            bg="#6f2dbd",
            fg="white",
            font=("Segoe UI", 24, "bold"),
        ).pack(pady=(20, 4))
        tk.Label(
            header,
            text="Entrada manual ou leitura automatica de orcamento com visual corporativo",
            bg="#6f2dbd",
            fg="#eadcff",
            font=("Segoe UI", 11),
        ).pack()

        nav = tk.Frame(self, bg="#eceff5", pady=12)
        nav.grid(row=1, column=0, sticky="ew")
        ttk.Button(nav, text="Tela inicial", style="Ghost.TButton", command=self.show_home).grid(row=0, column=0, padx=(20, 10))
        ttk.Button(nav, text="Digitar manualmente", style="Ghost.TButton", command=self.show_manual).grid(row=0, column=1, padx=10)
        ttk.Button(nav, text="Anexar orcamento", style="Ghost.TButton", command=self.show_attach).grid(row=0, column=2, padx=10)
        ttk.Button(nav, text="Exportar Excel", style="Primary.TButton", command=self.exportar_excel).grid(row=0, column=3, padx=10)

        self.content = tk.Frame(self, bg="#eceff5")
        self.content.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 10))
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(2, weight=1)

        self.status_var = tk.StringVar(value="Pronto para iniciar.")
        status = tk.Label(
            self,
            textvariable=self.status_var,
            anchor="w",
            bg="#dfe5f2",
            fg="#16253d",
            font=("Segoe UI", 10, "bold"),
            padx=14,
            pady=10,
        )
        status.grid(row=3, column=0, sticky="ew")

    def clear_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    def build_hero(self, parent, title, subtitle):
        hero = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        hero.grid(row=0, column=0, sticky="ew")
        hero.grid_columnconfigure(0, weight=1)
        tk.Label(hero, text=title, bg="#ffffff", fg="#10243e", font=("Segoe UI", 22, "bold")).grid(
            row=0, column=0, sticky="w", padx=24, pady=(22, 6)
        )
        tk.Label(hero, text=subtitle, bg="#ffffff", fg="#64748b", font=("Segoe UI", 11)).grid(
            row=1, column=0, sticky="w", padx=24, pady=(0, 22)
        )
        return hero

    def show_home(self):
        self.clear_content()
        self.content.grid_rowconfigure(1, weight=1)
        self.build_hero(
            self.content,
            "Escolha como deseja iniciar",
            "Use o modo manual para informar um item rapidamente ou anexe um orcamento para leitura automatica.",
        )

        cards = tk.Frame(self.content, bg="#eceff5")
        cards.grid(row=1, column=0, sticky="nsew", pady=(16, 0))
        cards.grid_columnconfigure((0, 1), weight=1)

        self._create_option_card(
            cards,
            0,
            "Modo manual",
            "Preencha Nome do Produto, NCM, IPI, ICMS, Preco, Quantidade, Frete e Filial. Ideal para cotacao rapida de um item.",
            "Abrir formulario",
            self.show_manual,
        )
        self._create_option_card(
            cards,
            1,
            "Anexar orcamento",
            "Selecione arquivos PDF, XLSX, XLSM, CSV, PNG, JPG ou JPEG. O sistema tenta identificar CNPJ, frete, filial e todos os itens automaticamente.",
            "Selecionar arquivo",
            self.show_attach,
        )
        self.status_var.set("Tela inicial carregada.")

    def _create_option_card(self, parent, col, title, desc, btn_text, command):
        card = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 8, 8 if col == 0 else 0))
        card.grid_columnconfigure(0, weight=1)
        tk.Label(card, text=title, bg="#ffffff", fg="#10243e", font=("Segoe UI", 18, "bold")).grid(
            row=0, column=0, sticky="w", padx=24, pady=(24, 10)
        )
        tk.Label(card, text=desc, bg="#ffffff", fg="#475569", font=("Segoe UI", 11), justify="left", wraplength=520).grid(
            row=1, column=0, sticky="w", padx=24
        )
        ttk.Button(card, text=btn_text, style="Primary.TButton", command=command).grid(
            row=2, column=0, sticky="w", padx=24, pady=24
        )

    def _build_psd_panel(self, parent, row):
        panel = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        panel.grid(row=row, column=0, sticky="ew", pady=(14, 0))
        panel.grid_columnconfigure((0, 1, 2), weight=1)

        tk.Label(
            panel,
            text="Resultado PSD da simulacao",
            bg="#ffffff",
            fg="#10243e",
            font=("Segoe UI", 15, "bold"),
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=22, pady=(16, 8))

        for idx, var in enumerate((self.psd_r_var, self.psd_s_var, self.psd_t_var)):
            box = tk.Label(
                panel,
                textvariable=var,
                bg="#f3e8ff",
                fg="#4a115f",
                font=("Segoe UI", 11, "bold"),
                padx=16,
                pady=18,
            )
            box.grid(
                row=1,
                column=idx,
                sticky="ew",
                padx=(22 if idx == 0 else 8, 22 if idx == 2 else 8),
                pady=(0, 16),
            )

    def _build_results_area(self, parent, row):
        box = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        box.grid(row=row, column=0, sticky="nsew", pady=(14, 0))
        box.grid_columnconfigure(0, weight=1)
        box.grid_rowconfigure(1, weight=1)

        top = tk.Frame(box, bg="#ffffff")
        top.grid(row=0, column=0, sticky="ew")
        top.grid_columnconfigure(1, weight=1)
        tk.Label(top, text="Resultados", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(
            row=0, column=0, sticky="w", padx=22, pady=14
        )
        self.item_combo_var = tk.StringVar()
        self.item_combo = ttk.Combobox(top, textvariable=self.item_combo_var, state="readonly")
        self.item_combo.grid(row=0, column=1, sticky="e", padx=22)
        self.item_combo.bind("<<ComboboxSelected>>", self.on_item_selected)

        table_frame = tk.Frame(box, bg="#ffffff")
        table_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        cols = ("idx", "codigo", "produto", "qtde", "ncm", "preco", "icms", "ipi", "frete", "r", "s", "t")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=11)
        headers = {
            "idx": "#",
            "codigo": "Codigo",
            "produto": "Produto",
            "qtde": "Qtde",
            "ncm": "NCM",
            "preco": "Preco Unitario",
            "icms": "ICMS",
            "ipi": "IPI",
            "frete": "Frete",
            "r": "R / Filial 4",
            "s": "S / Filial 4 Fora RN / Filial 2 Fora PE",
            "t": "T / Filial 3 e 5",
        }
        widths = {
            "idx": 45,
            "codigo": 90,
            "produto": 430,
            "qtde": 60,
            "ncm": 100,
            "preco": 115,
            "icms": 75,
            "ipi": 75,
            "frete": 70,
            "r": 135,
            "s": 230,
            "t": 155,
        }

        for col in cols:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor="center")
        self.tree.column("produto", anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.tag_configure("price_highlight", background="#fff4cc")
        self.tree.tag_configure("normal_row", background="#ffffff")

        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        xscroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

    def _mostrar_popup_resumo_tabela(self, titulo="Resumo da simulacao"):
        if not self.current_results:
            messagebox.showwarning(APP_TITLE, "Nao ha resultados para exibir.")
            return

        popup = tk.Toplevel(self)
        popup.title(titulo)
        popup.geometry("1180x620")
        popup.minsize(980, 500)
        popup.transient(self)
        popup.grab_set()
        popup.configure(bg="#f4f6fb")

        header = tk.Frame(popup, bg="#6f2dbd", height=70)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text=titulo,
            bg="#6f2dbd",
            fg="white",
            font=("Segoe UI", 17, "bold"),
        ).pack(side="left", padx=20, pady=16)

        info = tk.Frame(popup, bg="#f4f6fb")
        info.pack(fill="x", padx=16, pady=(12, 8))

        tk.Label(
            info,
            text=f"Itens processados: {len(self.current_results)}",
            bg="#f4f6fb",
            fg="#334155",
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left")

        table_wrap = tk.Frame(popup, bg="#f4f6fb")
        table_wrap.pack(fill="both", expand=True, padx=16, pady=(0, 12))
        table_wrap.grid_columnconfigure(0, weight=1)
        table_wrap.grid_rowconfigure(0, weight=1)

        cols = ("idx", "produto", "qtde", "ncm", "preco", "r", "s", "t")
        tree = ttk.Treeview(table_wrap, columns=cols, show="headings", height=14)

        headers = {
            "idx": "#",
            "produto": "Produto",
            "qtde": "Qtde",
            "ncm": "NCM",
            "preco": "Preco Unitario",
            "r": "PSD R / Filial 4",
            "s": "PSD S",
            "t": "PSD T / Filial 3 e 5",
        }
        widths = {
            "idx": 45,
            "produto": 380,
            "qtde": 60,
            "ncm": 100,
            "preco": 120,
            "r": 150,
            "s": 180,
            "t": 180,
        }

        for col in cols:
            tree.heading(col, text=headers[col])
            tree.column(col, width=widths[col], anchor="center")

        tree.column("produto", anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")

        yscroll = ttk.Scrollbar(table_wrap, orient="vertical", command=tree.yview)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll = ttk.Scrollbar(table_wrap, orient="horizontal", command=tree.xview)
        xscroll.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        for rec in self.current_results:
            tree.insert(
                "",
                "end",
                values=(
                    rec["idx"],
                    rec["produto"],
                    rec["qtde"],
                    rec["ncm"],
                    format_money(rec["preco"]),
                    format_money(rec["r"]),
                    format_money(rec["s"]),
                    format_money(rec["t"]),
                ),
            )

        footer = tk.Frame(popup, bg="#f4f6fb")
        footer.pack(fill="x", padx=16, pady=(0, 16))

        ttk.Button(footer, text="Fechar", style="Primary.TButton", command=popup.destroy).pack(side="right")

    def show_manual(self):
        self.clear_content()
        wrapper = tk.Frame(self.content, bg="#eceff5")
        wrapper.grid(row=0, column=0, sticky="nsew")
        wrapper.grid_columnconfigure(0, weight=1)
        wrapper.grid_rowconfigure(3, weight=1)

        self.build_hero(
            wrapper,
            "Entrada manual",
            "Preencha os campos abaixo e clique em Confirmar item manual para buscar o PSD com base na planilha.",
        )

        body = tk.Frame(wrapper, bg="#eceff5")
        body.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)

        left_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left_card.grid_columnconfigure(0, weight=1)

        top_actions = tk.Frame(left_card, bg="#ffffff")
        top_actions.grid(row=0, column=0, sticky="ew", padx=22, pady=(18, 8))
        top_actions.grid_columnconfigure(0, weight=1)
        tk.Label(top_actions, text="Formulario", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Button(top_actions, text="Limpar campos", style="Ghost.TButton", command=self.limpar_manual).grid(
            row=0, column=1, padx=(8, 8)
        )
        ttk.Button(
            top_actions,
            text="Confirmar item manual",
            style="Primary.TButton",
            command=self.confirmar_manual,
        ).grid(row=0, column=2)

        form = tk.Frame(left_card, bg="#ffffff")
        form.grid(row=1, column=0, sticky="ew", padx=22, pady=(0, 8))
        form.grid_columnconfigure(1, weight=1)

        self.manual_nome = tk.StringVar()
        self.manual_ncm = tk.StringVar()
        self.manual_ipi = tk.StringVar()
        self.manual_icms = tk.StringVar()
        self.manual_preco = tk.StringVar()
        self.manual_qtde = tk.StringVar(value="1")
        self.manual_frete = tk.StringVar(value="FOB")
        self.manual_filial = tk.StringVar(value=f"Natal - {CNPJ_RN}")

        self._add_entry(form, "Nome do Produto", self.manual_nome, 0)
        self._add_entry(form, "NCM", self.manual_ncm, 1)
        self._add_entry(form, "IPI (%)", self.manual_ipi, 2)
        self._add_entry(form, "ICMS (%)", self.manual_icms, 3)
        self._add_entry(form, "Preco Unitario", self.manual_preco, 4)
        self._add_entry(form, "Quantidade", self.manual_qtde, 5)

        tk.Label(form, text="Frete", bg="#ffffff", fg="#16253d", font=("Segoe UI", 11, "bold")).grid(
            row=6, column=0, sticky="w", pady=(10, 4)
        )
        frete_box = tk.Frame(form, bg="#ffffff")
        frete_box.grid(row=6, column=1, sticky="w", pady=(10, 4), padx=(12, 0))
        ttk.Radiobutton(frete_box, text="FOB", variable=self.manual_frete, value="FOB").pack(side="left", padx=(0, 16))
        ttk.Radiobutton(frete_box, text="CIF", variable=self.manual_frete, value="CIF").pack(side="left")

        tk.Label(form, text="Filial", bg="#ffffff", fg="#16253d", font=("Segoe UI", 11, "bold")).grid(
            row=7, column=0, sticky="w", pady=(12, 4)
        )
        filial_cb = ttk.Combobox(form, textvariable=self.manual_filial, state="readonly", width=54)
        filial_cb["values"] = [f"Natal - {CNPJ_RN}", f"Pernambuco - {CNPJ_PE}"]
        filial_cb.grid(row=7, column=1, sticky="ew", pady=(12, 4), padx=(12, 0))

        tk.Label(
            left_card,
            text="Campos IPI e ICMS iniciam em branco. O botao Limpar campos zera tudo e volta Qtde = 1.",
            bg="#fff4cc",
            fg="#5b4a00",
            font=("Segoe UI", 10),
            wraplength=720,
            padx=14,
            pady=12,
        ).grid(row=2, column=0, sticky="ew", padx=22, pady=(8, 18))

        right_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        right_card.grid(row=0, column=1, sticky="nsew")
        right_card.grid_columnconfigure(0, weight=1)
        tk.Label(right_card, text="Resumo do modo manual", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(
            row=0, column=0, sticky="w", padx=22, pady=(18, 10)
        )

        steps_text = "\n\n".join([
            "1. Informe Produto, NCM, Preco e Filial.",
            "2. Informe IPI e ICMS conforme a nota/orcamento.",
            "3. Escolha FOB ou CIF.",
            "4. Clique em Confirmar item manual.",
            "5. O sistema consulta a logica da planilha e retorna os PSDs R, S e T.",
        ])
        tk.Label(
            right_card,
            text=steps_text,
            justify="left",
            bg="#eef2ff",
            fg="#334155",
            font=("Segoe UI", 11),
            wraplength=420,
            padx=18,
            pady=18,
        ).grid(row=1, column=0, sticky="ew", padx=22)
        tk.Label(
            right_card,
            text="Se houver erro de NCM nao encontrado, a mensagem sera mostrada na tela.",
            justify="left",
            bg="#fff7e6",
            fg="#7c5a00",
            font=("Segoe UI", 10),
            wraplength=420,
            padx=16,
            pady=14,
        ).grid(row=2, column=0, sticky="ew", padx=22, pady=(18, 18))

        self._build_psd_panel(wrapper, row=2)
        self._build_results_area(wrapper, row=3)
        self.status_var.set("Modo manual ativo. Preencha os campos e clique em Confirmar item manual.")

    def _add_entry(self, parent, label, variable, row):
        tk.Label(parent, text=label, bg="#ffffff", fg="#16253d", font=("Segoe UI", 11, "bold")).grid(
            row=row, column=0, sticky="w", pady=(10, 4)
        )
        entry = ttk.Entry(parent, textvariable=variable, font=("Segoe UI", 11))
        entry.grid(row=row, column=1, sticky="ew", pady=(10, 4), padx=(12, 0))
        return entry

    def show_attach(self):
        self.clear_content()
        wrapper = tk.Frame(self.content, bg="#eceff5")
        wrapper.grid(row=0, column=0, sticky="nsew")
        wrapper.grid_columnconfigure(0, weight=1)
        wrapper.grid_rowconfigure(3, weight=1)

        self.build_hero(
            wrapper,
            "Leitura automatica de orcamento",
            "Anexe um arquivo e depois clique em Gerar simulacao para processar a precificacao.",
        )

        body = tk.Frame(wrapper, bg="#eceff5")
        body.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        body.grid_columnconfigure(0, weight=2)
        body.grid_columnconfigure(1, weight=3)

        left_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left_card.grid_columnconfigure(1, weight=1)

        top_actions = tk.Frame(left_card, bg="#ffffff")
        top_actions.grid(row=0, column=0, columnspan=2, sticky="ew", padx=22, pady=(18, 8))
        top_actions.grid_columnconfigure(0, weight=1)
        tk.Label(top_actions, text="Anexar orcamento", bg="#ffffff", fg="#10243e", font=("Segoe UI", 16, "bold")).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Checkbutton(top_actions, text="Modo aprendizado", variable=self.attach_learning_enabled).grid(row=0, column=1, padx=(8, 8))
        ttk.Button(top_actions, text="Ver diagnostico", style="Ghost.TButton", command=self.mostrar_diagnostico).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(top_actions, text="Gerar simulacao", style="Primary.TButton", command=self.processar_orcamento_anexado).grid(
            row=0, column=3
        )

        self.budget_path_var = tk.StringVar()
        ttk.Button(left_card, text="Selecionar arquivo", style="Ghost.TButton", command=self.open_budget).grid(
            row=1, column=0, padx=(22, 12), pady=(0, 12), sticky="w"
        )
        ttk.Entry(left_card, textvariable=self.budget_path_var, font=("Segoe UI", 10)).grid(
            row=1, column=1, padx=(0, 22), pady=(0, 12), sticky="ew"
        )

        self.detect_cnpj_var = tk.StringVar(value="CNPJ: -")
        self.detect_mode_var = tk.StringVar(value="Compra para: -")
        self.detect_frete_var = tk.StringVar(value="Frete: -")
        self.total_itens_var = tk.StringVar(value="Itens: 0")
        info_vars = [self.detect_cnpj_var, self.detect_mode_var, self.detect_frete_var, self.total_itens_var]
        for idx, var in enumerate(info_vars, start=2):
            tk.Label(
                left_card,
                textvariable=var,
                anchor="w",
                bg="#f6f8fc",
                fg="#16253d",
                font=("Segoe UI", 10, "bold"),
                padx=12,
                pady=10,
            ).grid(row=idx, column=0, columnspan=2, sticky="ew", padx=22, pady=(0, 8))

        tk.Label(
            left_card,
            text="Filial manual (usar quando o sistema nao identificar automaticamente)",
            bg="#ffffff",
            fg="#16253d",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=6, column=0, columnspan=2, sticky="w", padx=22, pady=(8, 4))

        attach_filial_cb = ttk.Combobox(
            left_card,
            textvariable=self.attach_filial_manual,
            state="readonly",
            width=46
        )
        attach_filial_cb["values"] = [
            f"Natal - {CNPJ_RN}",
            f"Pernambuco - {CNPJ_PE}",
        ]
        attach_filial_cb.grid(row=7, column=0, columnspan=2, sticky="ew", padx=22, pady=(0, 8))

        tk.Label(
            left_card,
            text="1. Selecione o arquivo.  2. Confira os dados detectados.  3. Se nao identificar RN/PE, escolha a filial acima.  4. Clique em Gerar simulacao.  5. Use Ver diagnostico para entender por que um item foi ou nao reconhecido.",
            bg="#fff4cc",
            fg="#5b4a00",
            font=("Segoe UI", 10),
            padx=12,
            pady=10,
            wraplength=920,
            justify="left",
        ).grid(row=8, column=0, columnspan=2, sticky="ew", padx=22, pady=(8, 18))

        right_card = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        right_card.grid(row=0, column=1, sticky="nsew")
        right_card.grid_columnconfigure(0, weight=1)
        tk.Label(
            right_card,
            text="Como funciona o preenchimento automatico",
            bg="#ffffff",
            fg="#10243e",
            font=("Segoe UI", 16, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=22, pady=(18, 10))

        proc = "\n\n".join([
            "1. Clique em Selecionar arquivo.",
            "2. Escolha o orcamento em PDF, Excel, CSV ou imagem.",
            "3. O sistema identifica CNPJ, tipo de frete e os itens.",
            "4. Clique em Gerar simulacao para calcular a precificacao.",
            "5. O sistema retorna os PSDs R, S e T de cada item.",
            "6. Ao final, exporte a planilha em Excel formatado.",
        ])
        tk.Label(
            right_card,
            text=proc,
            justify="left",
            bg="#eef2ff",
            fg="#334155",
            font=("Segoe UI", 11),
            wraplength=640,
            padx=18,
            pady=18,
        ).grid(row=1, column=0, sticky="ew", padx=22)

        model = "\n\n".join([
            "Cabecalhos recomendados no arquivo: Produto, NCM, IPI, ICMS, Frete, Preco, Quantidade e Codigo.",
            "Quanto mais estruturado vier o orcamento, melhor a leitura automatica.",
        ])
        tk.Label(
            right_card,
            text=model,
            justify="left",
            bg="#fff7e6",
            fg="#7c5a00",
            font=("Segoe UI", 10),
            wraplength=640,
            padx=16,
            pady=14,
        ).grid(row=2, column=0, sticky="ew", padx=22, pady=(18, 18))

        self._build_psd_panel(wrapper, row=2)
        self._build_results_area(wrapper, row=3)
        self.status_var.set("Modo de anexo ativo. Selecione o arquivo e clique em Gerar simulacao.")

    def limpar_manual(self):
        self.manual_nome.set("")
        self.manual_ncm.set("")
        self.manual_ipi.set("")
        self.manual_icms.set("")
        self.manual_preco.set("")
        self.manual_qtde.set("1")
        self.manual_frete.set("FOB")
        self.manual_filial.set(f"Natal - {CNPJ_RN}")
        self.items = []
        self.current_results = []
        self.psd_r_var.set("R / Filial 4: -")
        self.psd_s_var.set("S / Filial 4 Fora RN / Filial 2 Fora PE: -")
        self.psd_t_var.set("T / Filial 3 e 5: -")
        if hasattr(self, "tree"):
            for item_id in self.tree.get_children():
                self.tree.delete(item_id)
        if hasattr(self, "item_combo"):
            self.item_combo["values"] = []
            self.item_combo.set("")
        self.status_var.set("Campos do modo manual limpos.")

    def confirmar_manual(self):
        try:
            nome = self.manual_nome.get().strip()
            ncm = re.sub(r"\D", "", self.manual_ncm.get())
            icms_txt = self.manual_icms.get().strip()
            ipi_txt = self.manual_ipi.get().strip()

            if not nome:
                raise ValueError("Informe o Nome do Produto.")
            if not ncm:
                raise ValueError("Informe um NCM valido.")
            if not icms_txt:
                raise ValueError("Informe o ICMS (%).")
            if not ipi_txt:
                raise ValueError("Informe o IPI (%).")

            item = {
                "codigo": "MANUAL",
                "descricao": nome,
                "ncm": ncm,
                "ipi": parse_percent(ipi_txt),
                "icms": parse_percent(icms_txt),
                "frete": self.manual_frete.get(),
                "preco": to_decimal(self.manual_preco.get()),
                "qtde": int(to_decimal(self.manual_qtde.get()) or Decimal("1")),
            }
            if item["preco"] <= 0:
                raise ValueError("Informe um Preco Unitario maior que zero.")
            if item["qtde"] <= 0:
                item["qtde"] = 1

            self.items = [item]
            compra_para = "RN" if "Natal" in self.manual_filial.get() else "PE"
            self.detect_mode_var = tk.StringVar(value=f"Compra para: {compra_para}")
            self._populate_combo_labels()
            self.calcular_orcamento_inteiro()
            self._mostrar_popup_resumo_tabela("Resumo da simulacao manual")
            self.status_var.set("Item manual gerado com sucesso.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def mostrar_diagnostico(self):
        diagnostico = ""
        if self.budget_data and self.budget_data.get("diagnostics_text"):
            diagnostico = self.budget_data.get("diagnostics_text", "")
        else:
            diagnostico = self.reader.get_diagnostics_text()

        popup = tk.Toplevel(self)
        popup.title("Diagnóstico da leitura")
        popup.geometry("1080x680")
        popup.minsize(900, 520)
        popup.transient(self)
        popup.configure(bg="#f4f6fb")

        header = tk.Frame(popup, bg="#6f2dbd", height=68)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Diagnóstico da leitura e modo aprendizado", bg="#6f2dbd", fg="white", font=("Segoe UI", 16, "bold")).pack(side="left", padx=20, pady=16)

        top = tk.Frame(popup, bg="#f4f6fb")
        top.pack(fill="x", padx=16, pady=(12, 6))
        resumo = self.budget_data.get("learning_summary", self.reader.get_learning_summary()) if self.budget_data else self.reader.get_learning_summary()
        tk.Label(top, text=resumo, justify="left", anchor="w", bg="#eef2ff", fg="#334155", font=("Segoe UI", 10), padx=12, pady=10).pack(fill="x")

        body = tk.Frame(popup, bg="#f4f6fb")
        body.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        text_widget = tk.Text(body, wrap="word", font=("Consolas", 10), bg="white", fg="#10243e")
        text_widget.pack(side="left", fill="both", expand=True)
        yscroll = ttk.Scrollbar(body, orient="vertical", command=text_widget.yview)
        yscroll.pack(side="right", fill="y")
        text_widget.configure(yscrollcommand=yscroll.set)
        text_widget.insert("1.0", diagnostico or "Nenhum diagnóstico disponível.")
        text_widget.configure(state="disabled")

        footer = tk.Frame(popup, bg="#f4f6fb")
        footer.pack(fill="x", padx=16, pady=(0, 16))
        ttk.Button(footer, text="Fechar", style="Primary.TButton", command=popup.destroy).pack(side="right")

    def open_budget(self):
        path = filedialog.askopenfilename(
            title="Selecione o arquivo do orcamento",
            filetypes=[
                ("Arquivos suportados", "*.pdf *.xlsx *.xlsm *.csv *.png *.jpg *.jpeg"),
                ("PDF", "*.pdf"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Imagem", "*.png *.jpg *.jpeg"),
            ],
        )
        if not path:
            return
        try:
            data = self.reader.read(path, learning_enabled=self.attach_learning_enabled.get())
            self.budget_data = data
            self.budget_path_var.set(path)
            self.detect_cnpj_var.set(f"CNPJ: {data.get('cnpj') or 'nao identificado'}")
            self.detect_mode_var.set(f"Compra para: {data.get('compra_para') or 'nao identificada'}")
            self.detect_frete_var.set(f"Frete: {data.get('frete') or '-'}")
            self.total_itens_var.set(f"Itens: {len(data.get('items', []))}")
            estrategia = data.get("strategy", "-")
            fornecedor = data.get("supplier_name", "layout")
            self.status_var.set(f"Arquivo selecionado. Fornecedor/layout: {fornecedor} | estratégia: {estrategia}.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def processar_orcamento_anexado(self):
        try:
            if not self.budget_data:
                raise ValueError("Selecione primeiro um arquivo de orcamento.")

            self.items = self.budget_data.get("items", [])
            if not self.items:
                raise ValueError("Nao consegui identificar itens automaticamente nesse arquivo. Clique em 'Ver diagnostico' para entender o motivo e ajustar o layout/aprendizado.")

            compra_detectada = self.budget_data.get("compra_para", "")
            if compra_detectada not in ("RN", "PE"):
                compra_detectada = "RN" if "Natal" in self.attach_filial_manual.get() else "PE"

            if not hasattr(self, "detect_mode_var"):
                self.detect_mode_var = tk.StringVar(value="Compra para: -")
            self.detect_mode_var.set(f"Compra para: {compra_detectada}")

            if not hasattr(self, "detect_cnpj_var"):
                self.detect_cnpj_var = tk.StringVar(value="CNPJ: -")
            self.detect_cnpj_var.set(f"CNPJ: {CNPJ_RN if compra_detectada == 'RN' else CNPJ_PE}")

            self._populate_combo_labels()
            self.calcular_orcamento_inteiro()
            self._mostrar_popup_resumo_tabela("Resumo da simulacao do orcamento")

            self.status_var.set(f"Arquivo processado com sucesso. Itens encontrados: {len(self.items)}.")

        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def _populate_combo_labels(self):
        labels = []
        for i, item in enumerate(self.items, start=1):
            desc = item.get("descricao", "") or "(sem descricao)"
            labels.append(f"{i}. {item.get('codigo','')} | Qtde {item.get('qtde',1)} | {desc[:90]} | NCM: {item.get('ncm','')}")
        if hasattr(self, "item_combo"):
            self.item_combo["values"] = labels
            if labels:
                self.item_combo.current(0)

    def on_item_selected(self, event=None):
        idx = self.item_combo.current()
        if idx >= 0 and self.tree.get_children():
            self.tree.selection_set(self.tree.get_children()[idx])
            self.tree.see(self.tree.get_children()[idx])
            self._update_psd_panel_from_index(idx)

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        idx = int(values[0]) - 1
        if 0 <= idx < len(self.items):
            self.item_combo.current(idx)
            self._update_psd_panel_from_index(idx)

    def _update_psd_panel_from_index(self, idx):
        if 0 <= idx < len(self.current_results):
            rec = self.current_results[idx]
            self.psd_r_var.set(f"R / Filial 4: {format_money(rec['r'])}")
            self.psd_s_var.set(f"S / Filial 4 Fora RN / Filial 2 Fora PE: {format_money(rec['s'])}")
            self.psd_t_var.set(f"T / Filial 3 e 5: {format_money(rec['t'])}")

    def calcular_orcamento_inteiro(self):
        if not self.items:
            messagebox.showwarning(APP_TITLE, "Nao ha itens para calcular.")
            return

        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        self.current_results = []
        compra_para = ""
        cnpj_text = self.detect_mode_var.get().upper() if hasattr(self, "detect_mode_var") else ""

        if "RN" in cnpj_text:
            compra_para = "RN"
        elif "PE" in cnpj_text:
            compra_para = "PE"
        elif len(self.items) == 1 and hasattr(self, "manual_filial"):
            compra_para = "RN" if "Natal" in self.manual_filial.get() else "PE"

        if compra_para not in ("RN", "PE"):
            raise ValueError("Nao foi possivel identificar automaticamente RN ou PE.")

        for idx, item in enumerate(self.items, start=1):
            result = self.engine.calcular(
                compra_para,
                item["preco"],
                item["ncm"],
                item["icms"],
                item["ipi"],
                item["frete"],
            )

            record = {
                "idx": idx,
                "codigo": item.get("codigo", ""),
                "produto": item["descricao"],
                "qtde": item["qtde"],
                "ncm": item["ncm"],
                "preco": item["preco"],
                "icms": item["icms"],
                "ipi": item["ipi"],
                "frete": item["frete"],
                "r": result["r"],
                "s": result["s"],
                "t": result["t"],
            }
            self.current_results.append(record)

            self.tree.insert(
                "",
                "end",
                values=(
                    record["idx"],
                    record["codigo"],
                    record["produto"],
                    record["qtde"],
                    record["ncm"],
                    format_money(record["preco"]),
                    format_pct(record["icms"]),
                    format_pct(record["ipi"]),
                    record["frete"],
                    format_money(record["r"]),
                    format_money(record["s"]),
                    format_money(record["t"]),
                ),
                tags=("price_highlight",),
            )

        if self.current_results:
            self._update_psd_panel_from_index(0)
            filhos = self.tree.get_children()
            if filhos:
                self.tree.selection_set(filhos[0])
                self.tree.focus(filhos[0])
                self.tree.see(filhos[0])
            if hasattr(self, "item_combo") and self.item_combo["values"]:
                self.item_combo.current(0)

        self.update_idletasks()
        self.status_var.set(f"Precificacao concluida. Total de itens: {len(self.items)}.")

    def exportar_excel(self):
        if not self.current_results:
            messagebox.showwarning(APP_TITLE, "Calcule pelo menos um item antes de exportar.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Salvar resultado em Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="resultado_precificacao_profissional.xlsx",
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultado"
        headers = ["Item", "Codigo", "Produto", "Qtde", "NCM", "Preco Unitario", "ICMS", "IPI", "Frete", "Resultado R", "Resultado S", "Resultado T"]
        ws.append(headers)

        for rec in self.current_results:
            ws.append([
                rec["idx"],
                rec["codigo"],
                rec["produto"],
                rec["qtde"],
                rec["ncm"],
                float(q2(rec["preco"])),
                float(rec["icms"]),
                float(rec["ipi"]),
                rec["frete"],
                float(q2(rec["r"])),
                float(q2(rec["s"])),
                float(q2(rec["t"])),
            ])

        header_fill = PatternFill(fill_type="solid", fgColor="6F2DBD")
        header_font = Font(color="FFFFFF", bold=True)
        price_fill = PatternFill(fill_type="solid", fgColor="FFF4CC")
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        money_cols = {"F", "J", "K", "L"}
        pct_cols = {"G", "H"}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column_letter in money_cols and cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'
                elif cell.column_letter in pct_cols and cell.value is not None:
                    cell.number_format = '0.00%'
                if cell.column_letter == "F":
                    cell.fill = price_fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 70)

        ws.freeze_panes = "A2"
        wb.save(filepath)
        messagebox.showinfo(APP_TITLE, f"Arquivo exportado com sucesso:\n{filepath}")


if __name__ == "__main__":
    app = App()
    app.mainloop()