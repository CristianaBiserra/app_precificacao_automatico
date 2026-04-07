import os
import re
import csv
from decimal import Decimal, ROUND_HALF_UP
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_TITLE = "Assistente Profissional de Orçamento"
DEFAULT_WORKBOOK = os.path.join(os.path.dirname(os.path.abspath(__file__)), "1 NOVO SIMULADOR_PRECIFICACAO_V2.xlsx")

PIS_COFINS_RATE = Decimal("0.0925")
FRETE_FOB_RATE_RN = Decimal("0.10")
FRETE_FOB_RATE_PE = Decimal("0.05")
ROUND_MONEY = Decimal("0.01")

CNPJ_RN = "18.217.682/0004-05"
CNPJ_PE = "18.217.682/0001-54"
CNPJ_RN_DIGITS = "18217682000405"
CNPJ_PE_DIGITS = "18217682000154"

BG = "#f5f7fb"
CARD = "#ffffff"
PRIMARY = "#6f2dbd"
PRIMARY_DARK = "#59259a"
TEXT = "#1f2937"
MUTED = "#6b7280"
SUCCESS = "#0f766e"
SOFT = "#eef2ff"
WARNING = "#fff7d6"
BORDER = "#d8deea"


# ---------------------------
# Regras de cálculo
# ---------------------------
def q2(value: Decimal) -> Decimal:
    return value.quantize(ROUND_MONEY, rounding=ROUND_HALF_UP)


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip().replace("R$", "").replace("%", "").replace(" ", "")
    if text in ("-", "—", "None"):
        return Decimal("0")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return Decimal(text)


def parse_percent(value) -> Decimal:
    v = to_decimal(value)
    if v > Decimal("1"):
        v = v / Decimal("100")
    return v


def format_money(value: Decimal) -> str:
    value = q2(value)
    s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def format_pct(value: Decimal) -> str:
    v = (value * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}%"


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).strip().lower())


class PricingEngine:
    def __init__(self, workbook_path: str):
        self.ncm_factors = {}
        self.st_rates = {}
        self.load_workbook(workbook_path)

    def load_workbook(self, workbook_path: str):
        if not os.path.exists(workbook_path):
            raise FileNotFoundError(
                f"Planilha base não encontrada.\n\nColoque o arquivo '{os.path.basename(workbook_path)}' na mesma pasta do programa."
            )

        wb = openpyxl.load_workbook(workbook_path, data_only=False)

        ws_ncm = wb["Base Dados NCM"]
        self.ncm_factors.clear()
        for row in ws_ncm.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.ncm_factors[key] = {
                "fator_custo_filial_35": to_decimal(row[1]),
                "fator_psd_filial_35": to_decimal(row[3]),
                "fator_psd_filial_2_fora_pe": to_decimal(row[4]),
                "fator_psd_filial_4_dentro_rn": to_decimal(row[5]),
                "fator_psd_filial_4_fora_rn": to_decimal(row[6]),
            }

        ws_st = wb["Base ST RN"]
        self.st_rates.clear()
        for row in ws_st.iter_rows(min_row=2, values_only=True):
            ncm = row[0]
            if ncm in (None, ""):
                continue
            key = str(int(ncm)) if isinstance(ncm, (int, float)) else str(ncm).strip()
            self.st_rates[key] = to_decimal(row[1])

    def calcular(self, compra_para: str, preco_unitario: Decimal, ncm: str, icms: Decimal, ipi: Decimal, frete_tipo: str):
        ncm_key = str(ncm).strip()
        if ncm_key not in self.ncm_factors:
            raise KeyError(f"NCM {ncm_key} não encontrado na aba 'Base Dados NCM'.")

        fatores = self.ncm_factors[ncm_key]
        st_rate = self.st_rates.get(ncm_key, Decimal("0"))

        vlr_icms = preco_unitario * icms
        vlr_ipi = preco_unitario * ipi
        bc_pis_cofins = preco_unitario * (Decimal("1") - icms)
        pis_cofins = (bc_pis_cofins + vlr_ipi) * PIS_COFINS_RATE

        if compra_para == "RN":
            frete_fob = preco_unitario * FRETE_FOB_RATE_RN if frete_tipo == "FOB" else Decimal("0")
            vlr_st = preco_unitario * st_rate
            custo_filial_4 = (preco_unitario + vlr_ipi + vlr_st + frete_fob) - pis_cofins
            custo_filial_2 = custo_filial_4 * Decimal("0.88")
            custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
            r = custo_filial_4 * fatores["fator_psd_filial_4_dentro_rn"]
            s = custo_filial_4 * fatores["fator_psd_filial_4_fora_rn"]
            t = custo_filial_35 * fatores["fator_psd_filial_35"]
            return {"r": r, "s": s, "t": t}

        frete_fob = (preco_unitario + vlr_ipi) * FRETE_FOB_RATE_PE if frete_tipo == "FOB" else Decimal("0")
        custo_filial_2 = (preco_unitario + vlr_ipi + frete_fob) - pis_cofins - vlr_icms
        custo_filial_4 = custo_filial_2 * Decimal("1.12")
        custo_filial_35 = custo_filial_2 * fatores["fator_custo_filial_35"]
        r = custo_filial_4
        s = custo_filial_2 * fatores["fator_psd_filial_2_fora_pe"]
        t = custo_filial_35 * fatores["fator_psd_filial_35"]
        return {"r": r, "s": s, "t": t}


# ---------------------------
# Leitor universal
# ---------------------------
class UniversalBudgetReader:
    HEADER_ALIASES = {
        "descricao": ["descricao", "descrição", "produto", "item", "nomeproduto", "descricaoproduto", "material"],
        "ncm": ["ncm", "ncmsh", "codigoncm", "classfiscal", "classificacaofiscal", "classificaçãofiscal"],
        "ipi": ["ipi", "aliquotaipi", "percipi", "percentualipi"],
        "icms": ["icms", "aliquotaicms", "percicms", "percentualicms"],
        "frete": ["frete", "tipofrete", "modofrete", "fobcif", "ciffob"],
        "preco": ["preco", "preço", "precounitario", "preçounitário", "valorunitario", "vlunitario", "valor", "unitario", "unitário", "preco rsun"],
        "codigo": ["codigo", "código", "cod", "referencia", "referência"],
        "qtde": ["qtde", "quantidade", "qtd"],
    }

    def read(self, filepath: str):
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            return self._read_excel(filepath)
        if ext == ".csv":
            return self._read_csv(filepath)
        if ext == ".pdf":
            return self._read_pdf(filepath)
        if ext in (".png", ".jpg", ".jpeg"):
            return self._read_image(filepath)
        raise ValueError("Formato não suportado. Use PDF, XLS, XLSX, XLSM, CSV, PNG, JPG ou JPEG.")

    def _detect_cnpj_and_mode(self, text_dump: str):
        texto_numeros = re.sub(r"\D", "", text_dump)
        cnpjs = re.findall(r"\d{14}", texto_numeros)
        for cnpj in cnpjs:
            cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            if cnpj == CNPJ_RN_DIGITS or cnpj.endswith("000405"):
                return cnpj_formatado, "RN"
            if cnpj == CNPJ_PE_DIGITS or cnpj.endswith("000154"):
                return cnpj_formatado, "PE"
        upper = text_dump.upper()
        if "KGMLAN" in upper or "KGM LAN" in upper:
            return CNPJ_RN, "RN"
        return "", ""

    def _detect_frete(self, text_dump: str):
        upper = text_dump.upper()
        if "FRETE FOB" in upper or " FOB " in f" {upper} ":
            return "FOB"
        if "FRETE CIF" in upper or " CIF " in f" {upper} ":
            return "CIF"
        return "CIF"

    def _match_header(self, value):
        normalized = normalize_text(value)
        for field, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                if normalize_text(alias) == normalized:
                    return field
        return None

    def _dedupe_items(self, items):
        deduped = []
        seen = set()
        for item in items:
            key = (
                str(item.get("codigo", "")).strip(),
                str(item.get("descricao", "")).strip().upper(),
                re.sub(r"\D", "", str(item.get("ncm", ""))),
                str(q2(to_decimal(item.get("preco", "")))),
                str(item.get("qtde", "")),
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        return deduped

    def _post_process_items(self, items, default_frete, compra_para_override=""):
        processed = []
        for item in self._dedupe_items(items):
            desc = str(item.get("descricao", "")).strip()
            ncm = re.sub(r"\D", "", str(item.get("ncm", "")))
            preco = to_decimal(item.get("preco", ""))
            qtde_dec = to_decimal(item.get("qtde", "1"))
            qtde = int(qtde_dec) if qtde_dec > 0 else 1
            if qtde <= 0 or preco <= 0:
                continue
            if not desc and not ncm:
                continue
            processed.append({
                "codigo": str(item.get("codigo", "")).strip(),
                "descricao": desc,
                "ncm": ncm,
                "ipi": parse_percent(item.get("ipi", "")),
                "icms": parse_percent(item.get("icms", "")),
                "frete": self._detect_frete(str(item.get("frete", ""))) or default_frete,
                "preco": preco,
                "qtde": qtde,
                "compra_para": compra_para_override or str(item.get("compra_para", "")).upper(),
            })
        return processed

    def _extract_from_tabular_rows(self, rows, text_dump):
        found_items = []
        for idx, row in enumerate(rows[:30]):
            current_map = {}
            for col_idx, val in enumerate(row):
                if val in (None, ""):
                    continue
                matched = self._match_header(val)
                if matched:
                    current_map[matched] = col_idx
            if ("descricao" in current_map and "ncm" in current_map and "preco" in current_map):
                for data_row in rows[idx + 1:]:
                    if not any(v not in (None, "") for v in data_row):
                        continue
                    item = {}
                    for field, col_idx in current_map.items():
                        if col_idx < len(data_row):
                            item[field] = data_row[col_idx]
                    if "qtde" not in item:
                        item["qtde"] = 1
                    found_items.append(item)
                break
        if found_items:
            return found_items

        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        generic_items = []
        i = 0
        while i < len(lines):
            line = lines[i]
            m_head = re.match(r'^([A-Z0-9\-]+)\s+([\d\.,]+)\s+(\d+)\s+([\d\.,]+)(?:\s+[\d\.]+)?$', line)
            if m_head:
                codigo, preco, qtde, _total = m_head.groups()
                desc_lines = []
                j = i + 1
                while j < len(lines) and not re.search(r'\b\d{8}\b', lines[j]):
                    desc_lines.append(lines[j])
                    j += 1
                if j < len(lines):
                    tail = lines[j]
                    m_tail = re.search(r'(\d{8}).*?(\d{1,2}(?:[.,]\d{1,2})?).*?(\d{1,2}(?:[.,]\d{1,2})?)', tail)
                    if m_tail:
                        ncm, icms, ipi = m_tail.groups()
                        generic_items.append({
                            "codigo": codigo,
                            "descricao": " ".join(desc_lines).replace("#", "").strip(),
                            "ncm": ncm,
                            "icms": icms,
                            "ipi": ipi,
                            "preco": preco,
                            "frete": self._detect_frete(text_dump),
                            "qtde": qtde,
                        })
                        i = j
            i += 1
        return generic_items

    def _read_excel(self, filepath: str):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        text_parts, found_items = [], []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            sheet_text = []
            for row in rows:
                row_text = " ".join([str(v) for v in row if v not in (None, "")])
                if row_text:
                    sheet_text.append(row_text)
                    text_parts.append(row_text)
            found_items.extend(self._extract_from_tabular_rows(rows, "\n".join(sheet_text)))
        text_dump = "\n".join(text_parts)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete, compra_para)}

    def _read_csv(self, filepath: str):
        with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        text_dump = "\n".join([" ".join([str(v) for v in row if v not in (None, "")]) for row in rows])
        found_items = self._extract_from_tabular_rows(rows, text_dump)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete, compra_para)}

    def _read_pdf(self, filepath: str):
        try:
            import pdfplumber
        except Exception:
            raise ImportError("Para ler PDF, instale pdfplumber: python -m pip install pdfplumber")
        text_parts, found_items = [], []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                text_parts.append(text)
                tables = page.extract_tables() or []
                for table in tables:
                    if table:
                        found_items.extend(self._extract_from_tabular_rows(table, text))
        text_dump = "\n".join(text_parts)
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete, compra_para)}

    def _read_image(self, filepath: str):
        try:
            from PIL import Image
            import pytesseract
        except Exception:
            raise ImportError("Para ler imagem, instale Pillow e pytesseract, além do Tesseract OCR no Windows.")
        image = Image.open(filepath)
        text_dump = pytesseract.image_to_string(image, lang="por+eng", config="--psm 6")
        cnpj, compra_para = self._detect_cnpj_and_mode(text_dump)
        frete = self._detect_frete(text_dump)
        lines = [ln.strip() for ln in text_dump.splitlines() if ln.strip()]
        found_items = []
        current = {"codigo": "", "descricao": "", "ncm": "", "icms": "", "ipi": "", "preco": "", "frete": frete, "qtde": "1", "compra_para": compra_para}
        for line in lines:
            upper = line.upper()
            if not current["descricao"] and not any(x in upper for x in ["NCM", "ICMS", "IPI", "CNPJ", "R$"]):
                current["descricao"] = line
            if not current["ncm"]:
                m = re.search(r"\b(\d{8})\b", line)
                if m:
                    current["ncm"] = m.group(1)
            if not current["icms"] and "ICMS" in upper:
                p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                if p:
                    current["icms"] = p.group(1)
            if not current["ipi"] and "IPI" in upper:
                p = re.search(r'(\d{1,2}(?:[.,]\d{1,2})?)', line)
                if p:
                    current["ipi"] = p.group(1)
            if not current["preco"] and ("R$" in upper or "UNIT" in upper or "VALOR" in upper or "PRECO" in upper or "PREÇO" in upper):
                m = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})', line)
                if m:
                    current["preco"] = m.group(1)
            if current["qtde"] == "1":
                m = re.search(r'\bQTD(?:E)?\s*:?\s*(\d+)\b', upper)
                if m:
                    current["qtde"] = m.group(1)
        if current["descricao"] or current["ncm"]:
            found_items.append(current)
        return {"cnpj": cnpj, "compra_para": compra_para, "frete": frete, "items": self._post_process_items(found_items, frete, compra_para)}


class ModernCard(tk.Frame):
    def __init__(self, master, **kwargs):
        super().__init__(master, bg=CARD, highlightthickness=1, highlightbackground=BORDER, bd=0, **kwargs)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1440x900")
        self.minsize(1220, 760)
        self.configure(bg=BG)

        self.engine = PricingEngine(DEFAULT_WORKBOOK)
        self.reader = UniversalBudgetReader()
        self.items = []
        self.current_results = []
        self.current_mode = ""
        self.current_file = ""

        self.product_name_var = tk.StringVar()
        self.manual_ncm_var = tk.StringVar()
        self.manual_ipi_var = tk.StringVar(value="0")
        self.manual_icms_var = tk.StringVar(value="0")
        self.manual_preco_var = tk.StringVar(value="0")
        self.manual_qtde_var = tk.StringVar(value="1")
        self.manual_frete_var = tk.StringVar(value="CIF")
        self.manual_filial_var = tk.StringVar(value=f"Natal - {CNPJ_RN}")
        self.budget_path_var = tk.StringVar()
        self.detect_cnpj_var = tk.StringVar(value="CNPJ identificado: -")
        self.detect_mode_var = tk.StringVar(value="Compra para: -")
        self.detect_frete_var = tk.StringVar(value="Frete identificado: -")
        self.total_itens_var = tk.StringVar(value="Itens identificados: 0")
        self.status_var = tk.StringVar(value="Escolha uma das opções iniciais para começar.")

        self._build_styles()
        self._build_ui()
        self.show_home()

    def _build_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except Exception:
            pass
        style.configure("Primary.TButton", font=("Segoe UI", 11, "bold"), padding=(16, 12))
        style.configure("Secondary.TButton", font=("Segoe UI", 10), padding=(14, 10))
        style.configure("Clean.TEntry", padding=8)
        style.configure("Clean.TCombobox", padding=6)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview", rowheight=28, font=("Segoe UI", 10))

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header = tk.Frame(self, bg=PRIMARY, height=92)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)
        header.grid_columnconfigure(0, weight=1)

        tk.Label(
            header,
            text="ASSISTENTE PROFISSIONAL DE ORÇAMENTO",
            bg=PRIMARY,
            fg="white",
            font=("Segoe UI", 22, "bold")
        ).grid(row=0, column=0, pady=(18, 2))

        tk.Label(
            header,
            text="Entrada manual ou leitura automática de orçamento com visual corporativo",
            bg=PRIMARY,
            fg="#ece7ff",
            font=("Segoe UI", 10)
        ).grid(row=1, column=0)

        body = tk.Frame(self, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=18)
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(1, weight=1)

        toolbar = tk.Frame(body, bg=BG)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 12))

        self.btn_home = ttk.Button(toolbar, text="Tela inicial", command=self.show_home, style="Secondary.TButton")
        self.btn_home.pack(side="left")
        self.btn_manual = ttk.Button(toolbar, text="Digitar manualmente", command=self.show_manual_form, style="Secondary.TButton")
        self.btn_manual.pack(side="left", padx=8)
        self.btn_auto = ttk.Button(toolbar, text="Anexar orçamento", command=self.show_auto_mode, style="Secondary.TButton")
        self.btn_auto.pack(side="left")

        self.content = tk.Frame(body, bg=BG)
        self.content.grid(row=1, column=0, sticky="nsew")
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(0, weight=1)

        status = tk.Label(
            body,
            textvariable=self.status_var,
            anchor="w",
            bg="#eaeef8",
            fg=TEXT,
            font=("Segoe UI", 10, "bold"),
            padx=12,
            pady=10,
            relief="flat"
        )
        status.grid(row=2, column=0, sticky="ew", pady=(12, 0))

    def clear_content(self):
        for widget in self.content.winfo_children():
            widget.destroy()

    def show_home(self):
        self.current_mode = "home"
        self.clear_content()

        wrap = tk.Frame(self.content, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_columnconfigure((0, 1), weight=1)
        wrap.grid_rowconfigure(0, weight=1)

        left = ModernCard(wrap)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        right = ModernCard(wrap)
        right.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        self._build_choice_card(
            left,
            "01  |  Digitar as informações manualmente",
            "Abra um formulário guiado para informar produto, NCM, IPI, ICMS, preço, frete e filial da compra.",
            [
                "Nome do produto",
                "NCM",
                "IPI",
                "Frete FOB ou CIF",
                f"Filial Natal ({CNPJ_RN})",
                f"Filial Pernambuco ({CNPJ_PE})",
            ],
            "Abrir formulário manual",
            self.show_manual_form,
        )

        self._build_choice_card(
            right,
            "02  |  Anexar orçamento para preenchimento automático",
            "Selecione um orçamento em Excel, CSV, PDF ou imagem. O sistema tenta identificar CNPJ, frete, filial e itens automaticamente.",
            [
                "Anexar arquivo pelo seletor",
                "Leitura automática do cabeçalho e itens",
                "Detecção de CNPJ e filial RN/PE",
                "Listagem dos itens encontrados",
                "Cálculo em lote e exportação Excel",
            ],
            "Anexar orçamento",
            self.show_auto_mode,
        )

        self.status_var.set("Tela inicial carregada. Escolha a forma de entrada desejada.")

    def _build_choice_card(self, parent, title, desc, bullets, button_text, command):
        parent.grid_columnconfigure(0, weight=1)
        tk.Label(parent, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 17, "bold"), anchor="w").grid(row=0, column=0, sticky="w", padx=26, pady=(26, 8))
        tk.Label(parent, text=desc, bg=CARD, fg=MUTED, font=("Segoe UI", 10), justify="left", wraplength=520).grid(row=1, column=0, sticky="w", padx=26)

        bullets_frame = tk.Frame(parent, bg=SOFT)
        bullets_frame.grid(row=2, column=0, sticky="ew", padx=26, pady=22)
        bullets_frame.grid_columnconfigure(0, weight=1)
        for i, item in enumerate(bullets):
            tk.Label(bullets_frame, text=f"• {item}", bg=SOFT, fg=TEXT, font=("Segoe UI", 10), anchor="w").grid(row=i, column=0, sticky="ew", padx=16, pady=6)

        action = tk.Button(parent, text=button_text, command=command, bg=PRIMARY, fg="white", activebackground=PRIMARY_DARK,
                           activeforeground="white", relief="flat", font=("Segoe UI", 11, "bold"), cursor="hand2", padx=16, pady=12)
        action.grid(row=3, column=0, sticky="w", padx=26, pady=(0, 26))

    def show_manual_form(self):
        self.current_mode = "manual"
        self.clear_content()

        container = tk.Frame(self.content, bg=BG)
        container.pack(fill="both", expand=True)
        container.grid_columnconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=1)
        container.grid_rowconfigure(1, weight=1)

        top = ModernCard(container)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        tk.Label(top, text="Entrada manual", bg=CARD, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=24, pady=(20, 6))
        tk.Label(top, text="Preencha os campos abaixo. Após confirmar, o item será enviado para a tabela de resultados.", bg=CARD, fg=MUTED, font=("Segoe UI", 10)).pack(anchor="w", padx=24, pady=(0, 18))

        form_card = ModernCard(container)
        form_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        preview_card = ModernCard(container)
        preview_card.grid(row=1, column=1, sticky="nsew", padx=(8, 0))

        self._build_manual_form(form_card)
        self._build_manual_preview(preview_card)
        self.status_var.set("Modo manual ativo. Informe os dados do produto e clique em Confirmar item manual.")

    def _build_manual_form(self, parent):
        parent.grid_columnconfigure(1, weight=1)
        fields = [
            ("Nome do Produto", self.product_name_var),
            ("NCM", self.manual_ncm_var),
            ("IPI (%)", self.manual_ipi_var),
            ("ICMS (%)", self.manual_icms_var),
            ("Preço Unitário", self.manual_preco_var),
            ("Quantidade", self.manual_qtde_var),
        ]

        tk.Label(parent, text="Formulário", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=22, pady=(22, 18))

        for idx, (label_text, var) in enumerate(fields, start=1):
            tk.Label(parent, text=label_text, bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).grid(row=idx, column=0, sticky="w", padx=22, pady=8)
            ttk.Entry(parent, textvariable=var, style="Clean.TEntry").grid(row=idx, column=1, sticky="ew", padx=(0, 22), pady=8)

        frete_row = len(fields) + 1
        tk.Label(parent, text="FRETE", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).grid(row=frete_row, column=0, sticky="w", padx=22, pady=8)
        frete_wrap = tk.Frame(parent, bg=CARD)
        frete_wrap.grid(row=frete_row, column=1, sticky="w", padx=(0, 22), pady=8)
        ttk.Radiobutton(frete_wrap, text="FOB", variable=self.manual_frete_var, value="FOB").pack(side="left", padx=(0, 18))
        ttk.Radiobutton(frete_wrap, text="CIF", variable=self.manual_frete_var, value="CIF").pack(side="left")

        filial_row = frete_row + 1
        tk.Label(parent, text="Filial da compra", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).grid(row=filial_row, column=0, sticky="w", padx=22, pady=8)
        filial_combo = ttk.Combobox(
            parent,
            textvariable=self.manual_filial_var,
            state="readonly",
            style="Clean.TCombobox",
            values=[f"Natal - {CNPJ_RN}", f"Pernambuco - {CNPJ_PE}"]
        )
        filial_combo.grid(row=filial_row, column=1, sticky="ew", padx=(0, 22), pady=8)

        btn_frame = tk.Frame(parent, bg=CARD)
        btn_frame.grid(row=filial_row + 1, column=0, columnspan=2, sticky="w", padx=22, pady=(20, 22))

        tk.Button(btn_frame, text="Confirmar item manual", command=self.confirm_manual_item, bg=PRIMARY, fg="white",
                  activebackground=PRIMARY_DARK, activeforeground="white", relief="flat", font=("Segoe UI", 10, "bold"), padx=16, pady=10,
                  cursor="hand2").pack(side="left")
        tk.Button(btn_frame, text="Limpar campos", command=self.clear_manual_fields, bg="#e5e7eb", fg=TEXT,
                  relief="flat", font=("Segoe UI", 10), padx=16, pady=10, cursor="hand2").pack(side="left", padx=10)

    def _build_manual_preview(self, parent):
        tk.Label(parent, text="Resumo do modo manual", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=22, pady=(22, 10))

        summary = tk.Frame(parent, bg=SOFT)
        summary.pack(fill="x", padx=22, pady=(0, 18))

        textos = [
            "1. Digite as informações do produto.",
            "2. Escolha se o frete é FOB ou CIF.",
            f"3. Selecione a filial: Natal ({CNPJ_RN}) ou Pernambuco ({CNPJ_PE}).",
            "4. Clique em Confirmar item manual para enviar o item à grade.",
            "5. Exporte o resultado para Excel quando desejar.",
        ]
        for txt in textos:
            tk.Label(summary, text=txt, bg=SOFT, fg=TEXT, font=("Segoe UI", 10), anchor="w", justify="left").pack(fill="x", padx=16, pady=7)

        tips = tk.Frame(parent, bg=WARNING)
        tips.pack(fill="x", padx=22)
        tk.Label(
            tips,
            text="Campos recomendados para melhor cálculo: NCM, IPI, ICMS e Preço Unitário. A quantidade padrão é 1.",
            bg=WARNING,
            fg=TEXT,
            wraplength=500,
            justify="left",
            font=("Segoe UI", 10)
        ).pack(fill="x", padx=16, pady=14)

    def clear_manual_fields(self):
        self.product_name_var.set("")
        self.manual_ncm_var.set("")
        self.manual_ipi_var.set("0")
        self.manual_icms_var.set("0")
        self.manual_preco_var.set("0")
        self.manual_qtde_var.set("1")
        self.manual_frete_var.set("CIF")
        self.manual_filial_var.set(f"Natal - {CNPJ_RN}")
        self.status_var.set("Campos manuais limpos.")

    def confirm_manual_item(self):
        produto = self.product_name_var.get().strip()
        ncm = re.sub(r"\D", "", self.manual_ncm_var.get())
        if not produto:
            messagebox.showwarning(APP_TITLE, "Informe o nome do produto.")
            return
        if not ncm:
            messagebox.showwarning(APP_TITLE, "Informe um NCM válido.")
            return

        try:
            item = {
                "codigo": "MANUAL",
                "descricao": produto,
                "ncm": ncm,
                "ipi": parse_percent(self.manual_ipi_var.get()),
                "icms": parse_percent(self.manual_icms_var.get()),
                "frete": self.manual_frete_var.get() or "CIF",
                "preco": to_decimal(self.manual_preco_var.get()),
                "qtde": max(int(to_decimal(self.manual_qtde_var.get())), 1),
                "compra_para": "RN" if self.manual_filial_var.get().startswith("Natal") else "PE",
            }
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Erro ao validar dados manuais:\n{e}")
            return

        self.items = [item]
        self.current_file = "Entrada manual"
        self.budget_path_var.set(self.current_file)
        self.detect_cnpj_var.set(f"CNPJ identificado: {CNPJ_RN if item['compra_para'] == 'RN' else CNPJ_PE}")
        self.detect_mode_var.set(f"Compra para: {item['compra_para']}")
        self.detect_frete_var.set(f"Frete identificado: {item['frete']}")
        self.total_itens_var.set("Itens identificados: 1")
        self.show_results_screen(manual_mode=True)
        self.calcular_orcamento_inteiro(compra_para_override=item["compra_para"])
        self.status_var.set("Item manual processado com sucesso.")

    def show_auto_mode(self):
        self.current_mode = "auto"
        self.clear_content()

        wrapper = tk.Frame(self.content, bg=BG)
        wrapper.pack(fill="both", expand=True)
        wrapper.grid_columnconfigure(0, weight=1)
        wrapper.grid_rowconfigure(2, weight=1)

        intro = ModernCard(wrapper)
        intro.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        tk.Label(intro, text="Anexar orçamento para preenchimento automático", bg=CARD, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=24, pady=(20, 8))
        tk.Label(
            intro,
            text="Selecione um arquivo em Excel, CSV, PDF ou imagem. O sistema localizará cabeçalhos, tentará identificar CNPJ, frete e filial, e listará os itens encontrados.",
            bg=CARD, fg=MUTED, font=("Segoe UI", 10), wraplength=1150, justify="left"
        ).pack(anchor="w", padx=24, pady=(0, 18))

        attach = ModernCard(wrapper)
        attach.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        attach.grid_columnconfigure(1, weight=1)

        tk.Label(attach, text="Arquivo do orçamento", bg=CARD, fg=TEXT, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=(20, 8))
        ttk.Button(attach, text="Selecionar arquivo", command=self.open_budget, style="Primary.TButton").grid(row=1, column=0, padx=22, pady=(0, 18), sticky="w")
        ttk.Entry(attach, textvariable=self.budget_path_var, style="Clean.TEntry").grid(row=1, column=1, sticky="ew", padx=(0, 22), pady=(0, 18))

        info = tk.Frame(attach, bg=CARD)
        info.grid(row=2, column=0, columnspan=2, sticky="ew", padx=22, pady=(0, 22))
        info.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self._info_box(info, self.detect_cnpj_var, 0)
        self._info_box(info, self.detect_mode_var, 1)
        self._info_box(info, self.detect_frete_var, 2)
        self._info_box(info, self.total_itens_var, 3)

        tutorial = ModernCard(wrapper)
        tutorial.grid(row=2, column=0, sticky="nsew")
        tk.Label(tutorial, text="Passo a passo do preenchimento automático", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=24, pady=(22, 12))
        passos = [
            "1. Clique em Selecionar arquivo e escolha o orçamento.",
            "2. O sistema lê arquivos .xlsx, .xlsm, .xls, .csv, .pdf e imagens .png/.jpg/.jpeg.",
            "3. O processo procura colunas como Produto, NCM, IPI, ICMS, Preço, Qtde e Frete.",
            "4. Se encontrar CNPJ 18.217.682/0004-05, classifica a compra como RN. Se encontrar 18.217.682/0001-54, classifica como PE.",
            "5. Após a leitura, a tela de resultados é aberta com todos os itens identificados e precificados.",
            "6. Depois você pode exportar a tabela final para Excel formatado.",
        ]
        for passo in passos:
            tk.Label(tutorial, text=passo, bg=CARD, fg=TEXT, font=("Segoe UI", 10), anchor="w", justify="left", wraplength=1180).pack(fill="x", padx=24, pady=7)

        modelo = tk.Frame(tutorial, bg=SOFT)
        modelo.pack(fill="x", padx=24, pady=(18, 24))
        tk.Label(modelo, text="Modelo ideal de colunas do arquivo: Produto | NCM | IPI | ICMS | Preço | Qtde | Frete | Código", bg=SOFT, fg=TEXT, font=("Segoe UI", 10, "bold"), wraplength=1160, justify="left").pack(anchor="w", padx=16, pady=(14, 6))
        tk.Label(modelo, text="Quanto mais organizado estiver o orçamento, maior a chance de o preenchimento automático acertar todos os itens sem ajustes manuais.", bg=SOFT, fg=MUTED, font=("Segoe UI", 10), wraplength=1160, justify="left").pack(anchor="w", padx=16, pady=(0, 14))

        self.status_var.set("Modo automático ativo. Anexe um orçamento para iniciar a leitura.")

    def _info_box(self, parent, text_var, col):
        box = tk.Frame(parent, bg="#f4f6fb", highlightthickness=1, highlightbackground=BORDER)
        box.grid(row=0, column=col, sticky="ew", padx=6)
        tk.Label(box, textvariable=text_var, bg="#f4f6fb", fg=TEXT, font=("Segoe UI", 10, "bold"), anchor="w", padx=12, pady=12).pack(fill="both")

    def open_budget(self):
        path = filedialog.askopenfilename(
            title="Selecione o arquivo do orçamento",
            filetypes=[
                ("Arquivos suportados", "*.pdf *.xlsx *.xlsm *.xls *.csv *.png *.jpg *.jpeg"),
                ("PDF", "*.pdf"),
                ("Excel", "*.xlsx *.xlsm *.xls"),
                ("CSV", "*.csv"),
                ("Imagem", "*.png *.jpg *.jpeg"),
            ],
        )
        if not path:
            return
        try:
            data = self.reader.read(path)
            self.current_file = path
            self.budget_path_var.set(path)
            self.detect_cnpj_var.set(f"CNPJ identificado: {data.get('cnpj') or 'não identificado'}")
            self.detect_mode_var.set(f"Compra para: {data.get('compra_para') or 'não identificada'}")
            self.detect_frete_var.set(f"Frete identificado: {data.get('frete') or '-'}")
            self.items = data.get("items", [])

            if not self.items:
                raise ValueError("Não consegui identificar itens automaticamente nesse arquivo.")

            self.total_itens_var.set(f"Itens identificados: {len(self.items)}")
            self.show_results_screen(manual_mode=False)
            self.calcular_orcamento_inteiro(compra_para_override=data.get("compra_para", ""))
            self.status_var.set(f"Arquivo lido com sucesso. Itens únicos encontrados: {len(self.items)}.")

        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def show_results_screen(self, manual_mode=False):
        self.current_mode = "results"
        self.clear_content()

        main = tk.Frame(self.content, bg=BG)
        main.pack(fill="both", expand=True)
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(2, weight=1)

        top = ModernCard(main)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        top.grid_columnconfigure(1, weight=1)

        title_text = "Resultado da entrada manual" if manual_mode else "Resultado do orçamento anexado"
        tk.Label(top, text=title_text, bg=CARD, fg=TEXT, font=("Segoe UI", 18, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=22, pady=(20, 8))

        origem = self.current_file or self.budget_path_var.get() or "-"
        tk.Label(top, text="Origem dos dados:", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="w", padx=22)
        ttk.Entry(top, textvariable=tk.StringVar(value=origem), style="Clean.TEntry").grid(row=1, column=1, sticky="ew", padx=(0, 22), pady=(0, 12))

        info = tk.Frame(top, bg=CARD)
        info.grid(row=2, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 18))
        info.grid_columnconfigure((0, 1, 2, 3), weight=1)
        self._info_box(info, self.detect_cnpj_var, 0)
        self._info_box(info, self.detect_mode_var, 1)
        self._info_box(info, self.detect_frete_var, 2)
        self._info_box(info, self.total_itens_var, 3)

        actions = ModernCard(main)
        actions.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        actions.grid_columnconfigure(0, weight=1)

        left = tk.Frame(actions, bg=CARD)
        left.grid(row=0, column=0, sticky="w", padx=22, pady=18)
        tk.Label(left, text="Item selecionado:", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(side="left")
        self.item_combo_var = tk.StringVar()
        self.item_combo = ttk.Combobox(left, textvariable=self.item_combo_var, state="readonly", width=88, style="Clean.TCombobox")
        self.item_combo.pack(side="left", padx=(10, 0))
        self.item_combo.bind("<<ComboboxSelected>>", self.on_item_selected)

        labels = []
        for i, item in enumerate(self.items, start=1):
            desc = item.get("descricao", "") or "(sem descrição)"
            labels.append(f"{i}. {item.get('codigo', '')} | Qtde {item.get('qtde', 1)} | {desc[:90]} | NCM: {item.get('ncm', '')}")
        self.item_combo["values"] = labels
        if labels:
            self.item_combo.current(0)

        right = tk.Frame(actions, bg=CARD)
        right.grid(row=0, column=1, sticky="e", padx=22, pady=18)
        ttk.Button(right, text="Recalcular", command=self.calcular_orcamento_inteiro, style="Secondary.TButton").pack(side="left", padx=(0, 8))
        ttk.Button(right, text="Exportar Excel", command=self.exportar_excel, style="Primary.TButton").pack(side="left")

        table_box = ModernCard(main)
        table_box.grid(row=2, column=0, sticky="nsew")
        table_box.grid_columnconfigure(0, weight=1)
        table_box.grid_rowconfigure(1, weight=1)

        tk.Label(table_box, text="Todos os itens precificados", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=(20, 8))

        cols = ("idx", "codigo", "produto", "qtde", "ncm", "preco", "icms", "ipi", "frete", "r", "s", "t")
        self.tree = ttk.Treeview(table_box, columns=cols, show="headings", height=22)

        headers = {
            "idx": "#",
            "codigo": "Código",
            "produto": "Produto",
            "qtde": "Qtde",
            "ncm": "NCM",
            "preco": "Preço Unitário",
            "icms": "ICMS",
            "ipi": "IPI",
            "frete": "Frete",
            "r": "R / Filial 4",
            "s": "S / Filial 2",
            "t": "T / Filial 3 e 5",
        }
        widths = {
            "idx": 40,
            "codigo": 90,
            "produto": 660,
            "qtde": 60,
            "ncm": 100,
            "preco": 120,
            "icms": 75,
            "ipi": 75,
            "frete": 75,
            "r": 130,
            "s": 130,
            "t": 140,
        }

        for col in cols:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor="center")
        self.tree.column("produto", anchor="w")
        self.tree.grid(row=1, column=0, sticky="nsew", padx=(22, 0), pady=(0, 22))
        self.tree.tag_configure("price_highlight", background="#fff8db")
        self.tree.tag_configure("normal_row", background="#ffffff")
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        yscroll = ttk.Scrollbar(table_box, orient="vertical", command=self.tree.yview)
        yscroll.grid(row=1, column=1, sticky="ns", pady=(0, 22), padx=(0, 22))
        self.tree.configure(yscrollcommand=yscroll.set)

    def on_item_selected(self, event=None):
        idx = self.item_combo.current()
        if idx >= 0 and hasattr(self, "tree") and self.tree.get_children():
            self.tree.selection_set(self.tree.get_children()[idx])
            self.tree.see(self.tree.get_children()[idx])

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        idx = int(values[0]) - 1
        if hasattr(self, "item_combo") and 0 <= idx < len(self.items):
            self.item_combo.current(idx)

    def calcular_orcamento_inteiro(self, compra_para_override=""):
        if not self.items:
            messagebox.showwarning(APP_TITLE, "Selecione primeiro um arquivo ou informe um item manual.")
            return
        if not hasattr(self, "tree"):
            return

        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        self.current_results = []
        compra_para = (compra_para_override or "").upper()
        if compra_para not in ("RN", "PE"):
            mode_text = self.detect_mode_var.get().upper()
            if "RN" in mode_text:
                compra_para = "RN"
            elif "PE" in mode_text:
                compra_para = "PE"

        for idx, item in enumerate(self.items, start=1):
            item_compra_para = (item.get("compra_para") or compra_para).upper()
            if item_compra_para not in ("RN", "PE"):
                self.tree.insert(
                    "",
                    "end",
                    values=(idx, item.get("codigo", ""), "ERRO: filial não identificada para este item", item.get("qtde", 1), item.get("ncm", ""), "", "", "", item.get("frete", ""), "", "", ""),
                    tags=("normal_row",)
                )
                continue
            try:
                result = self.engine.calcular(
                    item_compra_para,
                    item["preco"],
                    item["ncm"],
                    item["icms"],
                    item["ipi"],
                    item["frete"],
                )
                record = {
                    "idx": idx,
                    "codigo": item.get("codigo", ""),
                    "produto": item.get("descricao", ""),
                    "qtde": item.get("qtde", 1),
                    "ncm": item.get("ncm", ""),
                    "preco": item["preco"],
                    "icms": item["icms"],
                    "ipi": item["ipi"],
                    "frete": item["frete"],
                    "r": result["r"],
                    "s": result["s"],
                    "t": result["t"],
                    "compra_para": item_compra_para,
                }
                self.current_results.append(record)
                self.tree.insert(
                    "",
                    "end",
                    values=(
                        record["idx"],
                        record["codigo"],
                        record["produto"],
                        record["qtde"],
                        record["ncm"],
                        format_money(record["preco"]),
                        format_pct(record["icms"]),
                        format_pct(record["ipi"]),
                        record["frete"],
                        format_money(record["r"]),
                        format_money(record["s"]),
                        format_money(record["t"]),
                    ),
                    tags=("price_highlight",)
                )
            except Exception as e:
                self.tree.insert(
                    "",
                    "end",
                    values=(idx, item.get("codigo", ""), f"ERRO: {str(e)}", item.get("qtde", 1), item.get("ncm", ""), "", "", "", item.get("frete", ""), "", "", ""),
                    tags=("normal_row",)
                )

        self.status_var.set(f"Itens processados: {len(self.items)}. Linhas calculadas com sucesso: {len(self.current_results)}.")

    def exportar_excel(self):
        if not self.current_results:
            messagebox.showwarning(APP_TITLE, "Calcule os itens antes de exportar.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Salvar resultado em Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="resultado_precificacao_profissional.xlsx",
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultado"

        headers = ["Item", "Codigo", "Produto", "Qtde", "NCM", "Preco Unitario", "ICMS", "IPI", "Frete", "Compra Para", "Resultado R", "Resultado S", "Resultado T"]
        ws.append(headers)

        for rec in self.current_results:
            ws.append([
                rec["idx"],
                rec["codigo"],
                rec["produto"],
                rec["qtde"],
                rec["ncm"],
                float(q2(rec["preco"])),
                float(rec["icms"]),
                float(rec["ipi"]),
                rec["frete"],
                rec["compra_para"],
                float(q2(rec["r"])),
                float(q2(rec["s"])),
                float(q2(rec["t"])),
            ])

        header_fill = PatternFill(fill_type="solid", fgColor="6F2DBD")
        header_font = Font(color="FFFFFF", bold=True)
        price_fill = PatternFill(fill_type="solid", fgColor="FFF8DB")
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        money_cols = {"F", "K", "L", "M"}
        pct_cols = {"G", "H"}

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column_letter in money_cols and cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'
                elif cell.column_letter in pct_cols and cell.value is not None:
                    cell.number_format = '0.00%'
                if cell.column_letter == "F":
                    cell.fill = price_fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 65)

        ws.freeze_panes = "A2"
        wb.save(filepath)
        messagebox.showinfo(APP_TITLE, f"Arquivo exportado com sucesso:\n{filepath}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
